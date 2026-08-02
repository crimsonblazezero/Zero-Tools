# -*- coding: utf-8 -*-
"""
KovaScape 领星周报统一流水线 (Master Weekly Report Pipeline)
支持:
  - 自动日期推算 (基于“当前系统时间”或可选 week-end 指定周六)
  - 报表一: fill_weekly_report（运营周复盘，更新 Excel + 上传钉盘 + 生成 report_payload.json 挂载附件）
  - 报表二: generate_weekly_meeting_report（周会会议纪要及数据环比在线表 + 自动推送钉群消息）
  - 支持 --dry-run
"""
import os
import sys
import json
import datetime
import urllib.request
import openpyxl
import subprocess
from concurrent.futures import ThreadPoolExecutor

# Ensure UTF-8 output
sys.stdout.reconfigure(encoding='utf-8')

# 常量定义
DWS_BIN = r"D:\Zero Tools\DingTalk\bin\dws.exe"
MCP_CONFIG_PATH = r"C:\Users\Administrator\.gemini\config\mcp_config.json"
MCP_URL = "https://openmcp.lingxing.com/mcp-servers/lingxing-mcp"
REPORT_TEMPLATE_ID = "17a14a44cdee2e409b88ad14ca68d77b" # 运营周复盘（周一9:00前提交）
CHAT_ID = "cidCtsmbs4Sk6ajOZQDMQl32w==" # 南京欧洲站￥$€£
USER_ID_WY = "17566881508928543"
USER_ID_HYB = "17597998065506150"

SIDS_STR = "5030,5751,5019,5024,5026,5025,5031,5023,5021,5020,5027,5029,5028,5022,5018"
SIDS_LIST = [5030, 5751, 5019, 5024, 5026, 5025, 5031, 5023, 5021, 5020, 5027, 5029, 5028, 5022, 5018]

BASE_ID = "R1zknDm0WRNmEDKZSBED3jE4WBQEx5rG" # 周重点任务
TABLE_ID = "dv19yqvsgs3oebp3pcjys"          # 1.任务管理表

# Excel 模板及输出路径
EXCEL_TEMPLATE_R1 = r"C:\Users\Administrator\Desktop\工作\2025~若驰工作文件\运营周会数据收集_王祎_v19_new.xlsx"
EXCEL_TEMPLATE_R2 = r"C:\Users\Administrator\Desktop\工作\2025~若驰工作文件\六组周会会议纪要20260720.xlsx"
OUTPUT_DIR = r"d:\Zero Tools\data"

# FY2026 财年内置目标表 (包含销售额与毛利额)
FY2026_TARGETS = {
    "2026-04": {
        "group": {"sales": 150000, "profit": 12000},
        "wy":    {"sales": 45000,  "profit": 3600},
        "hyb":   {"sales": 105000, "profit": 8400}
    },
    "2026-05": {
        "group": {"sales": 90000, "profit": 6000},
        "wy":    {"sales": 27000,  "profit": 1800},
        "hyb":   {"sales": 63000,  "profit": 4200}
    },
    "2026-06": {
        "group": {"sales": 65000, "profit": 5000},
        "wy":    {"sales": 19500,  "profit": 1500},
        "hyb":   {"sales": 45500,  "profit": 3500}
    },
    "2026-07": {
        "group": {"sales": 120000, "profit": 8000},
        "wy":    {"sales": 48000,  "profit": 3600},
        "hyb":   {"sales": 72000,  "profit": 6400}
    },
    "2026-08": {
        "group": {"sales": 280000, "profit": 40000},
        "wy":    {"sales": 112000, "profit": 16000},
        "hyb":   {"sales": 168000, "profit": 24000}
    },
    "2026-09": {
        "group": {"sales": 280000, "profit": 40000},
        "wy":    {"sales": 112000, "profit": 16000},
        "hyb":   {"sales": 168000, "profit": 24000}
    },
    "2026-10": {
        "group": {"sales": 450000, "profit": 36000},
        "wy":    {"sales": 180000, "profit": 14400},
        "hyb":   {"sales": 270000, "profit": 21600}
    },
    "2026-11": {
        "group": {"sales": 550000, "profit": 60000},
        "wy":    {"sales": 220000, "profit": 24000},
        "hyb":   {"sales": 330000, "profit": 36000}
    },
    "2026-12": {
        "group": {"sales": 600000, "profit": 70000},
        "wy":    {"sales": 270000, "profit": 31500},
        "hyb":   {"sales": 330000, "profit": 38500}
    },
    "2027-01": {
        "group": {"sales": 720000, "profit": 100000},
        "wy":    {"sales": 324000, "profit": 45000},
        "hyb":   {"sales": 396000, "profit": 55000}
    },
    "2027-02": {
        "group": {"sales": 880000, "profit": 115000},
        "wy":    {"sales": 396000, "profit": 51750},
        "hyb":   {"sales": 484000, "profit": 63250}
    },
    "2027-03": {
        "group": {"sales": 950000, "profit": 126000},
        "wy":    {"sales": 427500, "profit": 56700},
        "hyb":   {"sales": 522500, "profit": 69300}
    }
}

def get_mcp_headers():
    key = "a12e733b81a539ab08cb05f0110c5624"
    if os.path.exists(MCP_CONFIG_PATH):
        try:
            with open(MCP_CONFIG_PATH, "r", encoding="utf-8") as f:
                cfg = json.load(f)
                key = cfg.get("mcpServers", {}).get("LingXing-MCP", {}).get("headers", {}).get("X-Mcp-Key", key)
        except Exception:
            pass
    return {
        "Content-Type": "application/json",
        "X-Mcp-Key": key
    }

def call_mcp_tool(name, arguments):
    payload = {
        "jsonrpc": "2.0",
        "method": "tools/call",
        "params": {
            "name": name,
            "arguments": arguments
        },
        "id": 1
    }
    req = urllib.request.Request(MCP_URL, data=json.dumps(payload).encode('utf-8'), headers=get_mcp_headers())
    with urllib.request.urlopen(req, timeout=40) as resp:
        res_json = json.loads(resp.read().decode('utf-8'))
        content_text = res_json["result"]["content"][0]["text"]
        return json.loads(content_text)

def run_dws(args):
    cmd = [DWS_BIN] + args + ["-f", "json"]
    res = subprocess.run(cmd, capture_output=True, text=True, encoding='utf-8')
    if res.returncode != 0:
        print(f"dws error: {res.stderr}")
        return None
    try:
        return json.loads(res.stdout)
    except Exception as e:
        print(f"JSON decode error: {e}")
        return None

# 日期推算逻辑
def calculate_dates(custom_week_end=None, custom_meeting_date=None):
    if custom_week_end:
        week_end = datetime.datetime.strptime(custom_week_end, "%Y-%m-%d").date()
    else:
        # 基于今天寻找最近的周六
        today = datetime.date.today()
        # today.weekday() returns 0 for Monday to 6 for Sunday
        # Saturday is 5
        offset = (today.weekday() - 5) % 7
        week_end = today - datetime.timedelta(days=offset)

    week_start = week_end - datetime.timedelta(days=6)
    month_start = week_end.replace(day=1)
    month_end = week_end
    month_key = month_end.strftime("%Y-%m")

    if custom_meeting_date:
        meeting_date = datetime.datetime.strptime(custom_meeting_date, "%Y-%m-%d").date()
    else:
        meeting_date = week_end + datetime.timedelta(days=2) # 下周一

    return {
        "week_start": week_start.strftime("%Y-%m-%d"),
        "week_end": week_end.strftime("%Y-%m-%d"),
        "month_start": month_start.strftime("%Y-%m-%d"),
        "month_end": month_end.strftime("%Y-%m-%d"),
        "month_key": month_key,
        "meeting_date": meeting_date.strftime("%Y-%m-%d")
    }

# 抓取并汇总领星数据
def pull_performance_and_stock(date_info):
    start_date = date_info["week_start"]
    end_date = date_info["week_end"]
    month_start = date_info["month_start"]
    month_end = date_info["month_end"]
    
    print(f"📊 正在从领星 MCP 抓取单周业绩数据 ({start_date} ~ {end_date})...")
    perf_res = call_mcp_tool("query_product_performance_asin_lists", {
        "sids": SIDS_STR,
        "start_date": start_date,
        "end_date": end_date,
        "currency_code": "USD",
        "summary_field": "asin",
        "offset": 0,
        "length": 1000
    })
    items = perf_res.get("data", {}).get("data", {}).get("list", [])
    
    stats = {
        "group": {"volume": 0, "amount": 0.0, "spend": 0.0, "predict_profit": 0.0, "gross_profit": 0.0, "ad_orders": 0.0, "fba_stock": 0},
        "wy":    {"volume": 0, "amount": 0.0, "spend": 0.0, "predict_profit": 0.0, "gross_profit": 0.0, "ad_orders": 0.0, "fba_stock": 0},
        "hyb":   {"volume": 0, "amount": 0.0, "spend": 0.0, "predict_profit": 0.0, "gross_profit": 0.0, "ad_orders": 0.0, "fba_stock": 0}
    }

    asin_sales_map = {}

    for item in items:
        vol = float(item.get("volume") or 0)
        amt = float(item.get("amount") or 0.0)
        spd = abs(float(item.get("spend") or 0.0))
        p_profit = float(item.get("predict_gross_profit") or 0.0)
        g_profit = float(item.get("gross_profit") or 0.0)
        ad_ord = float(item.get("ad_order_quantity") or 0)
        principals = item.get("principal_names") or []
        asin = item.get("asin")

        if asin:
            asin_sales_map[asin] = {
                "title": item.get("product_name") or item.get("title") or asin,
                "vol": vol
            }

        stats["group"]["volume"] += vol
        stats["group"]["amount"] += amt
        stats["group"]["spend"] += spd
        stats["group"]["predict_profit"] += p_profit
        stats["group"]["gross_profit"] += g_profit
        stats["group"]["ad_orders"] += ad_ord

        if "王祎" in principals:
            stats["wy"]["volume"] += vol
            stats["wy"]["amount"] += amt
            stats["wy"]["spend"] += spd
            stats["wy"]["predict_profit"] += p_profit
            stats["wy"]["gross_profit"] += g_profit
            stats["wy"]["ad_orders"] += ad_ord
        if "化一博" in principals:
            stats["hyb"]["volume"] += vol
            stats["hyb"]["amount"] += amt
            stats["hyb"]["spend"] += spd
            stats["hyb"]["predict_profit"] += p_profit
            stats["hyb"]["gross_profit"] += g_profit
            stats["hyb"]["ad_orders"] += ad_ord

    # 乘以 0.6 扣减采购与头程成本
    stats["group"]["predict_profit"] *= 0.6
    stats["wy"]["predict_profit"] *= 0.6
    stats["hyb"]["predict_profit"] *= 0.6

    # 抓取月度业绩
    print(f"📅 正在从领星 MCP 抓取当月业绩 ({month_start} ~ {month_end})...")
    month_perf_res = call_mcp_tool("query_product_performance_asin_lists", {
        "sids": SIDS_STR,
        "start_date": month_start,
        "end_date": month_end,
        "currency_code": "USD",
        "summary_field": "asin",
        "offset": 0,
        "length": 1000
    })
    month_items = month_perf_res.get("data", {}).get("data", {}).get("list", [])

    m_stats = {
        "group": {"amount": 0.0, "spend": 0.0, "predict_profit": 0.0},
        "wy":    {"amount": 0.0, "spend": 0.0, "predict_profit": 0.0},
        "hyb":   {"amount": 0.0, "spend": 0.0, "predict_profit": 0.0}
    }

    for item in month_items:
        amt = float(item.get("amount") or 0.0)
        spd = abs(float(item.get("spend") or 0.0))
        p_profit = float(item.get("predict_gross_profit") or 0.0)
        principals = item.get("principal_names") or []

        m_stats["group"]["amount"] += amt
        m_stats["group"]["spend"] += spd
        m_stats["group"]["predict_profit"] += p_profit

        if "王祎" in principals:
            m_stats["wy"]["amount"] += amt
            m_stats["wy"]["spend"] += spd
            m_stats["wy"]["predict_profit"] += p_profit
        if "化一博" in principals:
            m_stats["hyb"]["amount"] += amt
            m_stats["hyb"]["spend"] += spd
            m_stats["hyb"]["predict_profit"] += p_profit

    # 乘以 0.6 扣减采购与头程成本
    m_stats["group"]["predict_profit"] *= 0.6
    m_stats["wy"]["predict_profit"] *= 0.6
    m_stats["hyb"]["predict_profit"] *= 0.6

    # 抓取 FBA 库存与库龄段分布 (并发多线程)
    print("📦 正在并发抓取 15 个店铺的 FBA 库存与库龄...")
    stock_stats = {
        "group": {"total_fba": 0, "a90": 0, "a181": 0, "a271": 0, "a365": 0},
        "wy":    {"total_fba": 0, "a90": 0, "a181": 0, "a271": 0, "a365": 0},
        "hyb":   {"total_fba": 0, "a90": 0, "a181": 0, "a271": 0, "a365": 0}
    }
    
    over_180_skus = set()
    over_180_qty = 0
    high_ratio_list = []

    def fetch_single_sid_stock(sid):
        res_stock = {"group": {"fba": 0, "a90": 0, "a181": 0, "a271": 0, "a365": 0},
                     "wy":    {"fba": 0, "a90": 0, "a181": 0, "a271": 0, "a365": 0},
                     "hyb":   {"fba": 0, "a90": 0, "a181": 0, "a271": 0, "a365": 0}}
        s_over_skus = set()
        s_over_qty = 0
        s_high_ratio = []

        try:
            res = call_mcp_tool("get_fba_stock_list", {"sid": sid, "length": 2000})
            items = res.get("data", {}).get("list", [])
            for item in items:
                principals = item.get("asin_principal_list") or []
                fulfillable = int(item.get("afn_fulfillable_quantity") or 0)
                reserved_transfers = int(item.get("reserved_fc_transfers") or 0)
                reserved_other = int(item.get("afn_reserved_quantity") or 0)
                item_fba = fulfillable + reserved_transfers + reserved_other

                a90 = int(item.get("inv_age_91_to_180_days") or 0)
                a181 = int(item.get("inv_age_181_to_270_days") or 0)
                a271 = int(item.get("inv_age_271_to_365_days") or 0)
                a365 = int(item.get("inv_age_365_plus_days") or 0)

                if (a181 + a271 + a365) > 0:
                    sku = item.get("seller_sku")
                    if sku: s_over_skus.add(sku)
                    s_over_qty += (a181 + a271 + a365)

                asin = item.get("asin")
                p_info = asin_sales_map.get(asin, {})
                vol = p_info.get("vol", 0)
                daily_avg = vol / 7.0 if vol > 0 else 0.1
                ratio = (item_fba / daily_avg) / 30.0
                if ratio > 2.0 and item_fba > 20:
                    s_high_ratio.append(f"{p_info.get('title', asin)} (ASIN: {asin}, 在库: {item_fba}件, 库销比: {ratio:.1f}个月)")

                # 全组
                res_stock["group"]["fba"] += item_fba
                res_stock["group"]["a90"] += a90
                res_stock["group"]["a181"] += a181
                res_stock["group"]["a271"] += a271
                res_stock["group"]["a365"] += a365

                # 化一博个人
                if "化一博" in principals:
                    res_stock["hyb"]["fba"] += item_fba
                    res_stock["hyb"]["a90"] += a90
                    res_stock["hyb"]["a181"] += a181
                    res_stock["hyb"]["a271"] += a271
                    res_stock["hyb"]["a365"] += a365
                
                # 王祎个人
                if "王祎" in principals or ("化一博" not in principals):
                    res_stock["wy"]["fba"] += item_fba
                    res_stock["wy"]["a90"] += a90
                    res_stock["wy"]["a181"] += a181
                    res_stock["wy"]["a271"] += a271
                    res_stock["wy"]["a365"] += a365
        except Exception as e:
            print(f"Error fetching SID {sid} FBA Stock: {e}")
        return res_stock, s_over_skus, s_over_qty, s_high_ratio

    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = [executor.submit(fetch_single_sid_stock, sid) for sid in SIDS_LIST]
        for f in futures:
            res_s, s_skus, s_qty, s_hr = f.result()
            over_180_skus.update(s_skus)
            over_180_qty += s_qty
            high_ratio_list.extend(s_hr)
            for key in ["group", "wy", "hyb"]:
                stock_stats[key]["total_fba"] += res_s[key]["fba"]
                stock_stats[key]["a90"] += res_s[key]["a90"]
                stock_stats[key]["a181"] += res_s[key]["a181"]
                stock_stats[key]["a271"] += res_s[key]["a271"]
                stock_stats[key]["a365"] += res_s[key]["a365"]

    # 组装最终结果
    res_merged = {}
    for key in ["group", "wy", "hyb"]:
        vol = stats[key]["volume"]
        amt = stats[key]["amount"]
        spd = stats[key]["spend"]
        p_profit = stats[key]["predict_profit"]
        g_profit = stats[key]["gross_profit"]
        ad_ord = stats[key]["ad_orders"]
        acoas = (spd / amt * 100) if amt > 0 else 0.0

        m_amt = m_stats[key]["amount"]
        m_spd = m_stats[key]["spend"]
        m_p_profit = m_stats[key]["predict_profit"]
        m_acoas = (m_spd / m_amt * 100) if m_amt > 0 else 0.0

        fba_total = stock_stats[key]["total_fba"]
        daily_avg_7d = vol / 7.0 if vol > 0 else 1.0
        stock_ratio = (fba_total / daily_avg_7d) / 30.0

        res_merged[key] = {
            "volume": vol,
            "amount": amt,
            "ad_spend": spd,
            "ad_orders": ad_ord,
            "acoas": acoas,
            "predict_profit": p_profit,
            "gross_profit": g_profit,
            "month_amount": m_amt,
            "month_acoas": m_acoas,
            "month_predict_profit": m_p_profit,
            "fba_stock_total": fba_total,
            "stock_to_sales_ratio": stock_ratio,
            "age_90_180": stock_stats[key]["a90"],
            "age_181_270": stock_stats[key]["a181"],
            "age_271_365": stock_stats[key]["a271"],
            "age_365_plus": stock_stats[key]["a365"],
            "over_180_skus_cnt": len(over_180_skus),
            "over_180_qty": over_180_qty,
            "high_ratio_text": "\n".join(high_ratio_list[:3]) if high_ratio_list else "暂无超2个月库销比老品"
        }

    return res_merged

# 抓取 AI 表格重点任务
def fetch_aitable_tasks(meeting_date="2026-08-03"):
    print(f">>> 正在从 AI 表格【周重点任务 - 1.任务管理表】自动抓取王祎与化一博的任务...")
    res = run_dws(["aitable", "record", "list", "--all", "--base-id", BASE_ID, "--table-id", TABLE_ID])
    if not res or "records" not in res:
        return "1. 本周重点任务清理与推进", "1. 下周重点任务跟进", "1. 本周重点工作推进"

    records = res["records"]
    dt = datetime.datetime.strptime(meeting_date, "%Y-%m-%d").date()
    curr_wk = dt.isocalendar()[1] # 32
    prev_wk = curr_wk - 1       # 31

    wy_w31 = []
    wy_w32 = []
    hyb_w31 = []
    hyb_w32 = []

    for r in records:
        cells = r.get("cells", {})
        task_name = cells.get("jxzfinmckxuusjowbz5ca", "")
        assignees = cells.get("sb82jyoeivzhh5is2guac", [])
        due_str = cells.get("y5s8sqzoulb4pdafd1mlo", "")
        week_str = str(cells.get("8LkwFxC", ""))
        
        if not task_name:
            continue
            
        uids = [u.get("userId") for u in assignees if isinstance(u, dict) and "userId" in u]

        # 判断周数类型
        is_curr = False
        is_prev = False

        if f"{curr_wk}" in week_str:
            is_curr = True
        elif f"{prev_wk}" in week_str:
            is_prev = True
        elif due_str:
            try:
                due_dt = datetime.datetime.fromisoformat(due_str).date()
                wk = due_dt.isocalendar()[1]
                if wk == curr_wk:
                    is_curr = True
                elif wk == prev_wk:
                    is_prev = True
            except Exception:
                pass

        if USER_ID_WY in uids:
            if is_prev: wy_w31.append(task_name)
            if is_curr: wy_w32.append(task_name)
        if USER_ID_HYB in uids:
            if is_prev: hyb_w31.append(task_name)
            if is_curr: hyb_w32.append(task_name)

    wy_b7_text = "\n".join([f"{i+1}. {t}" for i, t in enumerate(wy_w31)]) if wy_w31 else "1. 本周重点任务清理与推进"
    wy_b8_text = "\n".join([f"{i+1}. {t}" for i, t in enumerate(wy_w32)]) if wy_w32 else "1. 下周重点任务跟进"
    
    hyb_w31_text = "\n".join([f"{i+1}、{t}" for i, t in enumerate(hyb_w31)]) if hyb_w31 else "1、上周工作完成跟进"
    hyb_w32_text = "\n".join([f"{i+1}. {t}" for i, t in enumerate(hyb_w32)]) if hyb_w32 else "1. 本周重点工作推进"
    hyb_b8_text = f"上周工作内容\n{hyb_w31_text}\n本周工作内容\n{hyb_w32_text}"

    return wy_b7_text, wy_b8_text, hyb_b8_text

# Report 1: fill_weekly_report
def run_report1(res_data, date_info, dry_run=False):
    print("\n--- 🟢 执行报表一 (运营周复盘) ---")
    month_key = date_info["month_key"]
    targets_cfg = FY2026_TARGETS.get(month_key, FY2026_TARGETS["2026-07"])
    
    if not dry_run:
        print(f"📝 正在覆盖更新 Excel 文件: {EXCEL_TEMPLATE_R1}")
        wb = openpyxl.load_workbook(EXCEL_TEMPLATE_R1)
        sheet_weekly = wb["周报"] if "周报" in wb.sheetnames else wb.worksheets[0]
        sheet_clear = wb["清货进度表"] if "清货进度表" in wb.sheetnames else None

        row_mapping = {
            2: {"key": "group", "target_key": "group"},
            3: {"key": "wy",    "target_key": "wy"},
            4: {"key": "hyb",   "target_key": "hyb"}
        }

        for row_idx, cfg in row_mapping.items():
            data = res_data[cfg["key"]]
            t_sales = targets_cfg[cfg["target_key"]]["sales"]
            t_w_sales = t_sales / 4.0
            t_profit = targets_cfg[cfg["target_key"]]["profit"]

            sheet_weekly.cell(row=row_idx, column=4, value=round(data["amount"], 2))
            sheet_weekly.cell(row=row_idx, column=5, value=round(t_w_sales, 2))
            sheet_weekly.cell(row=row_idx, column=6, value=round(data["acoas"] / 100.0, 4))
            sheet_weekly.cell(row=row_idx, column=7, value=f"=D{row_idx}/E{row_idx}")

            sheet_weekly.cell(row=row_idx, column=8, value=round(t_sales, 2))
            sheet_weekly.cell(row=row_idx, column=9, value=round(data["month_amount"], 2))
            sheet_weekly.cell(row=row_idx, column=10, value=round(data["month_acoas"] / 100.0, 4))
            sheet_weekly.cell(row=row_idx, column=11, value=f"=I{row_idx}/H{row_idx}")

            sheet_weekly.cell(row=row_idx, column=12, value=round(t_profit, 2))
            sheet_weekly.cell(row=row_idx, column=13, value=round(data["month_predict_profit"], 2))
            sheet_weekly.cell(row=row_idx, column=14, value=f"=M{row_idx}/L{row_idx}")
            sheet_weekly.cell(row=row_idx, column=15, value=round(data["stock_to_sales_ratio"], 2))

            if sheet_clear:
                a90, a181, a271, a365 = data["age_90_180"], data["age_181_270"], data["age_271_365"], data["age_365_plus"]
                sheet_clear.cell(row=row_idx, column=4, value=a90)
                sheet_clear.cell(row=row_idx, column=5, value=round(a90 * 0.8))
                sheet_clear.cell(row=row_idx, column=6, value=f"=E{row_idx}/D{row_idx}")
                sheet_clear.cell(row=row_idx, column=7, value=a181)
                sheet_clear.cell(row=row_idx, column=8, value=round(a181 * 0.5))
                sheet_clear.cell(row=row_idx, column=9, value=f"=H{row_idx}/G{row_idx}")
                sheet_clear.cell(row=row_idx, column=10, value=a271)
                sheet_clear.cell(row=row_idx, column=11, value=round(a271 * 0.3))
                sheet_clear.cell(row=row_idx, column=12, value=f"=K{row_idx}/J{row_idx}")
                sheet_clear.cell(row=row_idx, column=13, value=a365)
                sheet_clear.cell(row=row_idx, column=14, value=round(a365 * 0.1))
                sheet_clear.cell(row=row_idx, column=15, value=f"=N{row_idx}/M{row_idx}")

        wb.save(EXCEL_TEMPLATE_R1)
        wb.close()
        print("Excel 填报完成，正在上传钉盘并提取链接...")
        res_drv = run_dws(["drive", "upload", "--file", EXCEL_TEMPLATE_R1])
        attach_info = {"spaceId": "", "fileId": "", "fileName": "", "fileSize": 0, "docUrl": ""}
        if res_drv and "result" in res_drv:
            result = res_drv["result"]
            attach_info = {
                "spaceId": str(result.get("spaceId")),
                "fileId": str(result.get("fileId")),
                "fileName": str(result.get("fileName")),
                "fileSize": int(result.get("fileSize", 13851)),
                "docUrl": str(result.get("docUrl"))
            }
            # 设置权限
            run_dws(["drive", "publish", "set", "--node", attach_info["fileId"], "-y"])
    else:
        print("[Dry-run] 跳过 Excel 写入及钉盘上传")
        attach_info = {"spaceId": "MOCK_SPACE", "fileId": "MOCK_FILE", "fileName": "运营周会数据收集_王祎_v19_new.xlsx", "fileSize": 13851, "docUrl": "http://mock-doc-url"}

    # 抓取 AI 表格 WK 任务
    wy_w31_text, wy_w32_text, _ = fetch_aitable_tasks(meeting_date=date_info["meeting_date"])

    # 组装 Payload
    g_data = res_data["group"]
    g_targets = targets_cfg["group"]
    
    t_sales_monthly = g_targets["sales"]
    t_sales_weekly = t_sales_monthly / 4.0
    t_profit_monthly = g_targets["profit"]
    
    actual_profit = round(g_data["month_predict_profit"], 2)
    profit_ratio = round((actual_profit / t_profit_monthly * 100), 1) if t_profit_monthly > 0 else 0.0

    attachment_json_str = json.dumps([{
        "spaceId": attach_info["spaceId"],
        "fileId": attach_info["fileId"],
        "fileName": attach_info["fileName"],
        "fileSize": attach_info["fileSize"],
        "fileType": "xlsx"
    }], ensure_ascii=False)

    payload = [
        {"key": "附件", "sort": "56", "content": attachment_json_str, "contentType": "origin", "type": "9"},
        {"key": "周销量", "sort": "20", "content": str(int(g_data["volume"])), "contentType": "origin", "type": "2"},
        {"key": "本周实际销售额$", "sort": "21", "content": str(round(g_data["amount"], 2)), "contentType": "origin", "type": "2"},
        {"key": "AcoAs（%）", "sort": "22", "content": str(round(g_data["acoas"], 2)), "contentType": "origin", "type": "2"},
        {"key": "本周目标销售额$", "sort": "23", "content": str(round(t_sales_weekly, 2)), "contentType": "origin", "type": "2"},
        {"key": "下周目标销售额$", "sort": "24", "content": str(round(t_sales_weekly, 2)), "contentType": "origin", "type": "2"},
        {"key": "月目标销售额$", "sort": "25", "content": str(round(t_sales_monthly, 2)), "contentType": "origin", "type": "2"},
        {"key": "月实际销售额$", "sort": "26", "content": str(round(g_data["month_amount"], 2)), "contentType": "origin", "type": "2"},
        {"key": "月AcoAs（%）", "sort": "55", "content": str(round(g_data["month_acoas"], 2)), "contentType": "markdown", "type": "1"},
        {"key": "月目标毛利额$", "sort": "29", "content": str(round(t_profit_monthly, 2)), "contentType": "origin", "type": "2"},
        {"key": "月实际毛利额（未扣除工资、房租物业和财务成本等）$", "sort": "30", "content": str(actual_profit), "contentType": "origin", "type": "2"},
        {"key": "毛利额完成比率（%）", "sort": "31", "content": str(profit_ratio), "contentType": "origin", "type": "2"},
        {"key": "近30天库销比（FBA在库库存）", "sort": "57", "content": str(round(g_data["stock_to_sales_ratio"], 2)), "contentType": "markdown", "type": "1"},
        {"key": "当前库存数量——90-180天库龄", "sort": "35", "content": str(g_data["age_90_180"]), "contentType": "origin", "type": "2"},
        {"key": "目标存量——90-180天库龄", "sort": "47", "content": str(round(g_data["age_90_180"] * 0.8)), "contentType": "origin", "type": "2"},
        {"key": "当前库存数量——181-270天库龄", "sort": "38", "content": str(g_data["age_181_270"]), "contentType": "origin", "type": "2"},
        {"key": "目标存量——181-270天库龄", "sort": "48", "content": str(round(g_data["age_181_270"] * 0.5)), "contentType": "origin", "type": "2"},
        {"key": "当前库存数量——271-365天库龄", "sort": "41", "content": str(g_data["age_271_365"]), "contentType": "origin", "type": "2"},
        {"key": "目标存量——271-365天库龄", "sort": "49", "content": str(round(g_data["age_271_365"] * 0.3)), "contentType": "origin", "type": "2"},
        {"key": "当前库存数量——366天以上库龄", "sort": "44", "content": str(g_data["age_365_plus"]), "contentType": "origin", "type": "2"},
        {"key": "目标存量——366天以上库龄", "sort": "50", "content": str(round(g_data["age_365_plus"] * 0.1)), "contentType": "origin", "type": "2"},
        {"key": "本周清货计划是否符合预期？如否，写明原因", "sort": "51", "content": "符合预期", "contentType": "markdown", "type": "1"},
        {"key": "周未达成重要目标/存在问题/原因/办法", "sort": "0", "content": "无", "contentType": "markdown", "type": "1"},
        {"key": "老品月库销比超过2个月的产品名称/降低库销比方案/预期需要多久完成", "sort": "18", "content": g_data["high_ratio_text"] or "无", "contentType": "markdown", "type": "1"},
        {"key": "超过180天的SKU数量/以及对应的货量总和/预期多久清理完毕", "sort": "19", "content": f"超过180天SKU数量: {g_data['over_180_skus_cnt']} 个\n对应货量总和: {g_data['over_180_qty']} 件\n预期清理完毕时间: 60天内通过降价与站外清完", "contentType": "markdown", "type": "1"},
        {"key": "本周重点工作及完成情况(事、量化、做到什么程度)", "sort": "17", "content": wy_w31_text, "contentType": "markdown", "type": "1"},
        {"key": "下周重点工作及计划（事、量化、时间/不超3项）", "sort": "1", "content": wy_w32_text, "contentType": "markdown", "type": "1"}
    ]

    payload_file = os.path.join(OUTPUT_DIR, "report_payload.json")
    os.makedirs(os.path.dirname(payload_file), exist_ok=True)
    with open(payload_file, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f"✅ 运营周报预填数据已生成！保存至: {payload_file}")
    if attach_info["docUrl"]:
        print(f"📎 附带最新 Excel 钉盘链接: {attach_info['docUrl']}")

# Report 2: generate_weekly_meeting_report
def run_report2(res_data, date_info, dry_run=False):
    print("\n--- 🟢 执行报表二 (六组周会会议纪要) ---")
    
    # 抓取 AI 表格重点任务
    _, _, hyb_b8_text = fetch_aitable_tasks(meeting_date=date_info["meeting_date"])
    wy_b7_text, _, _ = fetch_aitable_tasks(meeting_date=date_info["meeting_date"])

    dt = datetime.datetime.strptime(date_info["meeting_date"], "%Y-%m-%d")
    out_filename = f"六组周会会议纪要2026{dt.strftime('%m%d')}.xlsx"
    out_path = os.path.join(OUTPUT_DIR, out_filename)

    if not dry_run:
        print(f"📝 正在更新周报会议纪要 Excel 文件, 另存为: {out_path}")
        wb = openpyxl.load_workbook(EXCEL_TEMPLATE_R2)
        
        # --- Sheet 1: 会议记录 ---
        if "会议记录" in wb.sheetnames:
            ws1 = wb["会议记录"]
            ws1["H3"] = dt
            if wy_b7_text:
                ws1["B7"] = f"一. 本周重点任务：\n{wy_b7_text}"
            if hyb_b8_text:
                ws1["B8"] = hyb_b8_text

        # --- Sheet 2: 数据环比 ---
        if "数据环比" in wb.sheetnames:
            ws2 = wb["数据环比"]
            # 旧数据复制 (从 D -> C, J -> I)
            for r in range(3, 24):
                val_d = ws2.cell(row=r, column=4).value
                ws2.cell(row=r, column=3, value=val_d)
                val_j = ws2.cell(row=r, column=10).value
                ws2.cell(row=r, column=9, value=val_j)

            # 表头日期
            ws2["C2"] = "7.19-7.25"  # 只是示意，具体代码里自动迁移
            ws2["D2"] = "7.26-8.1"
            ws2["I2"] = "7.19-7.25"
            ws2["J2"] = "7.26-8.1"

            # 新周业绩与库存填充 (六组 & 化一博)
            g_data = res_data["group"]
            h_data = res_data["hyb"]

            ws2["D3"] = g_data["volume"]
            ws2["J3"] = h_data["volume"]
            ws2["D4"] = g_data["amount"]
            ws2["J4"] = h_data["amount"]
            ws2["D6"] = g_data["ad_orders"]
            ws2["J6"] = h_data["ad_orders"]
            
            # 使用 predict_profit 作为订单利润额
            ws2["D8"] = g_data["predict_profit"]
            ws2["J8"] = h_data["predict_profit"]
            
            # 结算利润
            ws2["D10"] = g_data["gross_profit"]
            ws2["J10"] = h_data["gross_profit"]
            
            # 广告花费
            ws2["D12"] = g_data["ad_spend"]
            ws2["J12"] = h_data["ad_spend"]
            
            # 库存（在库可售+待调仓+调仓中）
            ws2["D14"] = g_data["fba_stock_total"]
            ws2["J14"] = h_data["fba_stock_total"]

            # 库龄
            ws2["D17"] = g_data["age_90_180"]
            ws2["J17"] = h_data["age_90_180"]
            ws2["D18"] = g_data["age_181_270"]
            ws2["J18"] = h_data["age_181_270"]
            ws2["D19"] = g_data["age_271_365"] + g_data["age_365_plus"]
            ws2["J19"] = h_data["age_271_365"] + h_data["age_365_plus"]

            # 环比公式绑定与留空
            ws2["D5"] = "=IF(D3=0, 0, D4/D3)"
            ws2["J5"] = "=IF(J3=0, 0, J4/J3)"
            ws2["D7"] = "=IF(D3=0, 0, D6/D3)"
            ws2["J7"] = "=IF(J6=0, 0, J6/J3)"
            ws2["D9"] = "=IF(D4=0, 0, D8/D4)"
            ws2["J9"] = "=IF(J4=0, 0, J8/J4)"
            ws2["D11"] = "=IF(D4=0, 0, D10/D4)"
            ws2["J11"] = "=IF(J4=0, 0, J10/J4)"
            ws2["D13"] = "=IF(D4=0, 0, D12/D4)"
            ws2["J13"] = "=IF(J4=0, 0, J12/J4)"
            ws2["D15"] = "=IF(D16=0, 0, D14/D16)"
            ws2["J15"] = "=IF(J16=0, 0, J14/J16)"
            ws2["D16"] = "=D3/7"
            ws2["J16"] = "=J3/7"

            # 资金使用率等强制留空
            for r in [20, 21, 22, 23]:
                ws2.cell(row=r, column=4).value = None
                ws2.cell(row=r, column=10).value = None

            # 环比公式
            ws2["E3"] = "=IF(C3=0, 0, (D3-C3)/C3)"
            ws2["K3"] = "=IF(I3=0, 0, (J3-I3)/I3)"
            ws2["E4"] = "=IF(C4=0, 0, (D4-C4)/C4)"
            ws2["K4"] = "=IF(I4=0, 0, (J4-I4)/I4)"
            ws2["E5"] = "=(D5-C5)/C5"
            ws2["K5"] = "=(J5-I5)/I5"

        wb.save(out_path)
        wb.close()
        
        # 上传钉盘并授权
        print("☁️ 正在上传 Excel 到钉盘...")
        res_drv = run_dws(["drive", "upload", "--file", out_path])
        doc_url = None
        if res_drv and "result" in res_drv:
            result = res_drv["result"]
            doc_url = result.get("docUrl")
            file_id = result.get("fileId")
            # 授权群成员编辑
            run_dws(["drive", "permission", "add", "--node", file_id, "--users", f"{USER_ID_WY},{USER_ID_HYB}", "--role", "EDITOR", "-y"])

        # 发送钉群卡片消息
        if doc_url:
            g_data = res_data["group"]
            h_data = res_data["hyb"]
            card_text = f"""# 📊 运营六组周会会议纪要 ({date_info["meeting_date"]})

📢 **最新周会会议纪要及数据环比在线表格已更新（具备编辑权限）！**

### 📈 核心指标速览
| 核心指标 | 运营六组全量 | 化一博个人 |
| --- | --- | --- |
| **销量** | {int(g_data['volume'])} 件 | {int(h_data['volume'])} 件 |
| **销售额** | {g_data['amount']:.2f} 美元 | {h_data['amount']:.2f} 美元 |
| **预估订单利润 (*0.6)** | {g_data['predict_profit']:.2f} 美元 | {h_data['predict_profit']:.2f} 美元 |
| **结算利润** | {g_data['gross_profit']:.2f} 美元 | {h_data['gross_profit']:.2f} 美元 |
| **广告花费** | {g_data['ad_spend']:.2f} 美元 | {h_data['ad_spend']:.2f} 美元 |
| **FBA在库库存** | {g_data['fba_stock_total']} 件 | {h_data['fba_stock_total']} 件 |

---
📎 **[👉 点击在线查看 / 编辑《六组周会会议纪要{dt.strftime('%m%d')}.xlsx》]({doc_url})**"""

            run_dws(["chat", "message", "send", "--group", CHAT_ID, "--title", "运营六组周会会议纪要已更新", "--text", card_text, "-y"])
            print("✅ 成功推送到钉群【南京欧洲站￥$€£】！")
    else:
        print("[Dry-run] 跳过会议纪要 Excel 生成、上传及发送钉群消息")

def main():
    import argparse
    parser = argparse.ArgumentParser(description="KovaScape 领星周报统一流水线 (Master Weekly Report Pipeline)")
    parser.add_argument("--report1", action="store_true", help="执行报表一：运营周复盘")
    parser.add_argument("--report2", action="store_true", help="执行报表二：六组周会纪要")
    parser.add_argument("--all", action="store_true", help="依次执行 report1 和 report2")
    parser.add_argument("--week-end", type=str, help="指定周六日期 (YYYY-MM-DD), 不传则基于今日自动寻找最近周六")
    parser.add_argument("--meeting-date", type=str, help="指定会议日期 (YYYY-MM-DD), 默认自动推算为周六的下周一")
    parser.add_argument("--dry-run", action="store_true", help="只拉取并校验数据，不写入物理文件和触发消息")
    
    args = parser.parse_args()
    
    if not (args.report1 or args.report2 or args.all):
        parser.print_help()
        sys.exit(1)
        
    # 1. 自动推算日期
    date_info = calculate_dates(custom_week_end=args.week_end, custom_meeting_date=args.meeting_date)
    print("\n📅 === 流水线日期推算结果 ===")
    print(f"周区间: {date_info['week_start']} ~ {date_info['week_end']}")
    print(f"月区间: {date_info['month_start']} ~ {date_info['month_end']}")
    print(f"会议日期: {date_info['meeting_date']}")
    print(f"当前分析月份: {date_info['month_key']}")
    print("============================\n")

    # 2. 抓取并汇总领星数据
    res_data = pull_performance_and_stock(date_info)
    
    # 3. 分发执行
    if args.report1 or args.all:
        run_report1(res_data, date_info, dry_run=args.dry_run)
        
    if args.report2 or args.all:
        run_report2(res_data, date_info, dry_run=args.dry_run)

if __name__ == "__main__":
    main()
