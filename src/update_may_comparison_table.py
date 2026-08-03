# -*- coding: utf-8 -*-
"""
更新《王祎-当月销售分析-202606_Filled_v7.xlsx》中【二、核心指标环比分析】5月份的最新核算数据
"""
import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

OUTPUT_DIR = r"d:\Zero Tools\data"

p3 = os.path.join(OUTPUT_DIR, '王祎-当月销售分析-202606_Filled_v7.xlsx')
wb = openpyxl.load_workbook(p3, data_only=False)

if '王祎-当月销售分析' in wb.sheetnames:
    ws = wb['王祎-当月销售分析']
    # 查找【二、核心指标环比分析】数据表位置
    # Row 21 (假设) 是“二、核心指标环比分析”
    print("=== 开始检索并更新【二、核心指标环比分析】 ===")
    for r in range(15, ws.max_row + 1):
        v1 = str(ws.cell(row=r, column=2).value or '').strip()
        if '利润额' in v1:
            # Col C (2026年5月) | Col D (2026年6月) | Col E (环比)
            # 更新 2026年5月 利润额为 14907.14 (全量最新校准数)
            ws.cell(row=r, column=3).value = 14907.14
            print(f"✅ Row {r} (利润额) Col C 已更新为: 14907.14")
        elif '利润率' in v1:
            ws.cell(row=r, column=3).value = 0.2148
            print(f"✅ Row {r} (利润率) Col C 已更新为: 21.48%")
        elif '销售额' in v1 and '广告' not in v1:
            ws.cell(row=r, column=3).value = 69388.75
            print(f"✅ Row {r} (销售额) Col C 已更新为: 69388.75")

    wb.save(p3)
    wb.close()
    print("\n🎉 《王祎-当月销售分析-202606_Filled_v7.xlsx》环比表已成功刷新！")
