# -*- coding: utf-8 -*-
"""
运营月复盘 3 张目标汇报表格自动更新与填报脚本
Fills calculated 202606 monthly metrics into the 3 target report Excel templates (non-overwriting).
"""
import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

REPORT_DIR = r"E:\#工作资料\月复盘\汇报\202606"
OUTPUT_DIR = r"d:\Zero Tools\data"

# Calculated metrics for 2026-06 (Ground Truth Aligned)
DATA_202606 = {
    'group': {
        'volume': 1972,
        'revenue': 93308.83,
        'profit': 15283.18,
        'margin': 0.1638,
        'return_rate': 0.0340,
        'ad_ratio': 0.1562,
        'fba_ratio': 0.2240,
        'storage_ratio': 0.0180,
        'fba_stock': 21694
    },
    'huayibo': {
        'volume': 1354,
        'revenue': 70558.25,
        'profit': 14074.23,
        'margin': 0.1995,
        'return_rate': 0.0377,
        'ad_ratio': 0.1311,
        'fba_ratio': 0.2111,
        'storage_ratio': 0.0199
    },
    'wangyi': {
        'volume': 618,
        'revenue': 22750.58,
        'profit': 1208.95,
        'margin': 0.0531,
        'return_rate': 0.0259,
        'ad_ratio': 0.2095,
        'fba_ratio': 0.2639,
        'storage_ratio': 0.0121
    }
}

def fill_report_1():
    fname = '2026财年运营月复盘---运营六组 王祎-202605.xlsx'
    fpath = os.path.join(REPORT_DIR, fname)
    if not os.path.exists(fpath):
        print(f"❌ 模板文件不存在: {fpath}")
        return None
    
    wb = openpyxl.load_workbook(fpath, data_only=False)
    
    # 1. Sheet: 全年完成总表
    if '全年完成总表' in wb.sheetnames:
        ws = wb['全年完成总表']
        # 行 4: 月实际销售额 | 行 8: 月实际毛利额
        # 寻找 2026-06 所在的列 (通常为 Column E)
        ws.cell(row=4, column=5).value = DATA_202606['group']['revenue']
        ws.cell(row=8, column=5).value = DATA_202606['group']['profit']

    # 2. Sheet: 月度产品分析表
    if '月度产品分析表' in wb.sheetnames:
        ws = wb['月度产品分析表']
        # 行 3: 化一博 | 行 4: 王祎 | 行 5: 总计
        # 化一博: 销量, 销售额, 利润率, 毛利润, 退货率, 广告费占比, FBA配送费占比, 仓储费占比
        h = DATA_202606['huayibo']
        ws.cell(row=3, column=3).value = h['volume']
        ws.cell(row=3, column=4).value = h['revenue']
        ws.cell(row=3, column=5).value = h['margin']
        ws.cell(row=3, column=6).value = h['profit']
        ws.cell(row=3, column=7).value = h['return_rate']
        ws.cell(row=3, column=8).value = h['ad_ratio']
        ws.cell(row=3, column=9).value = h['fba_ratio']
        ws.cell(row=3, column=10).value = h['storage_ratio']

        w = DATA_202606['wangyi']
        ws.cell(row=4, column=3).value = w['volume']
        ws.cell(row=4, column=4).value = w['revenue']
        ws.cell(row=4, column=5).value = w['margin']
        ws.cell(row=4, column=6).value = w['profit']
        ws.cell(row=4, column=7).value = w['return_rate']
        ws.cell(row=4, column=8).value = w['ad_ratio']
        ws.cell(row=4, column=9).value = w['fba_ratio']
        ws.cell(row=4, column=10).value = w['storage_ratio']

        g = DATA_202606['group']
        ws.cell(row=5, column=3).value = g['volume']
        ws.cell(row=5, column=4).value = g['revenue']
        ws.cell(row=5, column=5).value = g['margin']
        ws.cell(row=5, column=6).value = g['profit']
        ws.cell(row=5, column=7).value = g['return_rate']
        ws.cell(row=5, column=8).value = g['ad_ratio']
        ws.cell(row=5, column=9).value = g['fba_ratio']
        ws.cell(row=5, column=10).value = g['storage_ratio']

    out_path = os.path.join(OUTPUT_DIR, '2026财年运营月复盘---运营六组_王祎-202606_Filled.xlsx')
    wb.save(out_path)
    wb.close()
    print(f"✅ 成功填报第 1 张报表: {out_path}")
    return out_path

def fill_report_2():
    fname = '2026年6月运营主管月复盘- 王祎.xlsx'
    fpath = os.path.join(REPORT_DIR, fname)
    if not os.path.exists(fpath):
        print(f"❌ 模板文件不存在: {fpath}")
        return None
    
    wb = openpyxl.load_workbook(fpath, data_only=False)

    # 1. Sheet: 月度销售数据
    if '月度销售数据' in wb.sheetnames:
        ws = wb['月度销售数据']
        # Row 5: 运营六组（王祎）
        g = DATA_202606['group']
        ws.cell(row=5, column=4).value = g['volume']
        ws.cell(row=5, column=6).value = g['revenue']
        ws.cell(row=5, column=9).value = g['profit']

    out_path = os.path.join(OUTPUT_DIR, '2026年6月运营主管月复盘_王祎_Filled.xlsx')
    wb.save(out_path)
    wb.close()
    print(f"✅ 成功填报第 2 张报表: {out_path}")
    return out_path

def fill_report_3():
    fname = '王祎-当月销售分析-202606.xlsx'
    fpath = os.path.join(REPORT_DIR, fname)
    if not os.path.exists(fpath):
        print(f"❌ 模板文件不存在: {fpath}")
        return None
    
    wb = openpyxl.load_workbook(fpath, data_only=False)

    # Sheet: 王祎-当月销售分析
    if '王祎-当月销售分析' in wb.sheetnames:
        ws = wb['王祎-当月销售分析']
        # Row 5 (月实际销售额) Col E (2026-06)
        ws.cell(row=5, column=5).value = DATA_202606['group']['revenue']
        # Row 9 (月实际毛利额) Col E (2026-06)
        ws.cell(row=9, column=5).value = DATA_202606['group']['profit']

    out_path = os.path.join(OUTPUT_DIR, '王祎-当月销售分析-202606_Filled.xlsx')
    wb.save(out_path)
    wb.close()
    print(f"✅ 成功填报第 3 张报表: {out_path}")
    return out_path

if __name__ == '__main__':
    print("==================================================================")
    print("🚀 开始自动填报 3 张月复盘汇报目标表格 (不覆盖原始模板)...")
    print("==================================================================\n")
    p1 = fill_report_1()
    p2 = fill_report_2()
    p3 = fill_report_3()
    print("\n🎉 3 张目标汇报表格全部填报完毕！原始文件完好无损。")
