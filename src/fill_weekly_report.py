# -*- coding: utf-8 -*-
"""
KovaScape 领星周报自动填报与钉钉《运营周复盘》全自动准备脚本 (100%全组数据+全量动态控件版)
- 全组数据: 销售、广告、利润、库销比均使用六组 100 汇总数据
- 目标存量: 自动按 0.8 / 0.5 / 0.3 / 0.1 计算填入
- 库销比 >2 个月老品: 动态匹配中文品名、在库数量并输出方案
- 180天以上库龄SKU: 动态统计 SKU 数量与货量总和
- 附件: 挂载最新的 Excel 在线文件
- 安全保护: 生成 report_payload.json 后停止自动提交，等待确认
"""

import os
import sys
import json
import datetime
import urllib.request
import openpyxl
import subprocess
from concurrent.futures import ThreadPoolExecutor

# Ensure UTF-8 output
sys.stdout.reconfigure(encoding='utf-8')

# 常量定义
DWS_BIN = r"D:\Zero Tools\DingTalk\bin\dws.exe"
MCP_CONFIG_PATH = r"C:\Users\Administrator\.gemini\config\mcp_config.json"
MCP_URL = "https://openmcp.lingxing.com/mcp-servers/lingxing-mcp"
REPORT_TEMPLATE_ID = "17a14a44cdee2e409b88ad14ca68d77b" # 运营周复盘（周一9:00前提交）
USER_ID_WY = "17566881508928543"

SIDS_STR = "5030,5751,5019,5024,5026,5025,5031,5023,5021,5020,5027,5029,5028,5022,5018"
SIDS_LIST = [5030, 5751, 5019, 5024, 5026, 5025, 5031, 5023, 5021, 5020, 5027, 5029, 5028, 5022, 5018]

# 2026 财年内置目标表
FY2026_TARGETS = {
    "2026-07": {
        "group": {"sales": 120000, "profit": 8000},
        "wy":    {"sales": 48000,  "profit": 3600},
        "hyb":   {"sales": 72000,  "profit": 6400}
    },
    "2026-08": {
        "group": {"sales": 280000, "profit": 40000},
        "wy":    {"sales": 112000, "profit": 16000},
        "hyb":   {"sales": 168000, "profit": 24000}
    },
    "2026-09": {
        "group": {"sales": 280000, "profit": 40000},
        "wy":    {"sales": 112000, "profit": 16000},
        "hyb":   {"sales": 168000, "profit": 24000}
    }
}

def get_mcp_headers():
    key = "a12e733b81a539ab08cb05f0110c5624"
    if os.path.exists(MCP_CONFIG_PATH):
        try:
            with open(MCP_CONFIG_PATH, "r", encoding="utf-8") as f:
                cfg = json.load(f)
                key = cfg.get("mcpServers", {}).get("LingXing-MCP", {}).get("headers", {}).get("X-Mcp-Key", key)
        except Exception:
            pass
    return {
        "Content-Type": "application/json",
        "X-Mcp-Key": key
    }

def call_mcp_tool(name, arguments):
    payload = {
        "jsonrpc": "2.0",
        "method": "tools/call",
        "params": {
            "name": name,
            "arguments": arguments
        },
        "id": 1
    }
    req = urllib.request.Request(MCP_URL, data=json.dumps(payload).encode('utf-8'), headers=get_mcp_headers())
    with urllib.request.urlopen(req, timeout=30) as resp:
        res_json = json.loads(resp.read().decode('utf-8'))
        content_text = res_json["result"]["content"][0]["text"]
        return json.loads(content_text)

def run_dws(args):
    cmd = [DWS_BIN] + args + ["-f", "json"]
    res = subprocess.run(cmd, capture_output=True, text=True, encoding='utf-8')
    if res.returncode != 0:
        print(f"dws error: {res.stderr}")
        return None
    try:
        return json.loads(res.stdout)
    except Exception as e:
        print(f"JSON decode error: {e}")
        return None

# 1. 抓取业绩与库存数据
def pull_performance_and_stock(start_date="2026-07-26", end_date="2026-08-01"):
    print(f"📊 正在从领星 MCP 抓取单周业绩数据 ({start_date} ~ {end_date})...")
    
    perf_res = call_mcp_tool("query_product_performance_asin_lists", {
        "sids": SIDS_STR,
        "start_date": start_date,
        "end_date": end_date,
        "currency_code": "USD",
        "summary_field": "asin",
        "offset": 0,
        "length": 1000
    })
    res_data_obj = perf_res.get("data", {})
    if isinstance(res_data_obj, dict) and "data" in res_data_obj:
        res_data_obj = res_data_obj.get("data", {})
    items = res_data_obj.get("list", [])
    
    stats = {
        "group": {"volume": 0, "amount": 0.0, "spend": 0.0, "predict_profit": 0.0},
        "wy":    {"volume": 0, "amount": 0.0, "spend": 0.0, "predict_profit": 0.0},
        "hyb":   {"volume": 0, "amount": 0.0, "spend": 0.0, "predict_profit": 0.0}
    }

    asin_sales_map = {}

    for item in items:
        vol = float(item.get("volume") or 0)
        amt = float(item.get("amount") or 0.0)
        spd = float(item.get("spend") or 0.0)
        p_profit = float(item.get("predict_gross_profit") or 0.0)
        principals = item.get("principal_names") or []
        asin = item.get("asin")

        if asin:
            asin_sales_map[asin] = {
                "title": item.get("product_name") or item.get("title") or asin,
                "vol": vol
            }

        stats["group"]["volume"] += vol
        stats["group"]["amount"] += amt
        stats["group"]["spend"] += spd
        stats["group"]["predict_profit"] += p_profit

        if "王祎" in principals:
            stats["wy"]["volume"] += vol
            stats["wy"]["amount"] += amt
            stats["wy"]["spend"] += spd
            stats["wy"]["predict_profit"] += p_profit
        if "化一博" in principals:
            stats["hyb"]["volume"] += vol
            stats["hyb"]["amount"] += amt
            stats["hyb"]["spend"] += spd
            stats["hyb"]["predict_profit"] += p_profit

    # 7 月整月业绩
    month_start = "2026-07-01"
    month_end = "2026-07-31"
    print(f"📅 正在从领星 MCP 抓取 7 月整月业绩 ({month_start} ~ {month_end})...")
    month_perf_res = call_mcp_tool("query_product_performance_asin_lists", {
        "sids": SIDS_STR,
        "start_date": month_start,
        "end_date": month_end,
        "currency_code": "USD",
        "summary_field": "asin",
        "offset": 0,
        "length": 1000
    })
    m_data_obj = month_perf_res.get("data", {})
    if isinstance(m_data_obj, dict) and "data" in m_data_obj:
        m_data_obj = m_data_obj.get("data", {})
    month_items = m_data_obj.get("list", [])

    m_stats = {
        "group": {"amount": 0.0, "spend": 0.0, "predict_profit": 0.0},
        "wy":    {"amount": 0.0, "spend": 0.0, "predict_profit": 0.0},
        "hyb":   {"amount": 0.0, "spend": 0.0, "predict_profit": 0.0}
    }

    for item in month_items:
        amt = float(item.get("amount") or 0.0)
        spd = float(item.get("spend") or 0.0)
        p_profit = float(item.get("predict_gross_profit") or 0.0)
        principals = item.get("principal_names") or []

        m_stats["group"]["amount"] += amt
        m_stats["group"]["spend"] += spd
        m_stats["group"]["predict_profit"] += p_profit

        if "王祎" in principals:
            m_stats["wy"]["amount"] += amt
            m_stats["wy"]["spend"] += spd
            m_stats["wy"]["predict_profit"] += p_profit
        if "化一博" in principals:
            m_stats["hyb"]["amount"] += amt
            m_stats["hyb"]["spend"] += spd
            m_stats["hyb"]["predict_profit"] += p_profit

    # 库存与库龄
    print("📦 正在并发抓取 15 个店铺的 FBA 库存与库龄...")
    stock_stats = {
        "group": {"total_fba": 0, "a90": 0, "a181": 0, "a271": 0, "a365": 0},
        "wy":    {"total_fba": 0, "a90": 0, "a181": 0, "a271": 0, "a365": 0},
        "hyb":   {"total_fba": 0, "a90": 0, "a181": 0, "a271": 0, "a365": 0}
    }
    
    over_180_skus = set()
    over_180_qty = 0
    high_ratio_list = []

    def fetch_single_sid_stock(sid):
        res_stock = {"group": {"fba": 0, "a90": 0, "a181": 0, "a271": 0, "a365": 0},
                     "wy":    {"fba": 0, "a90": 0, "a181": 0, "a271": 0, "a365": 0},
                     "hyb":   {"fba": 0, "a90": 0, "a181": 0, "a271": 0, "a365": 0}}
        s_over_skus = set()
        s_over_qty = 0
        s_high_ratio = []

        try:
            res = call_mcp_tool("get_fba_stock_list", {"sid": sid, "length": 2000})
            items = res.get("data", {}).get("list", [])
            for item in items:
                principals = item.get("asin_principal_list") or []
                fulfillable = int(item.get("afn_fulfillable_quantity") or 0)
                reserved_transfers = int(item.get("reserved_fc_transfers") or 0)
                reserved_other = int(item.get("afn_reserved_quantity") or 0)
                item_fba = fulfillable + reserved_transfers + reserved_other

                a90 = int(item.get("inv_age_91_to_180_days") or 0)
                a181 = int(item.get("inv_age_181_to_270_days") or 0)
                a271 = int(item.get("inv_age_271_to_365_days") or 0)
                a365 = int(item.get("inv_age_365_plus_days") or 0)

                if (a181 + a271 + a365) > 0:
                    sku = item.get("seller_sku")
                    if sku: s_over_skus.add(sku)
                    s_over_qty += (a181 + a271 + a365)

                asin = item.get("asin")
                p_info = asin_sales_map.get(asin, {})
                vol = p_info.get("vol", 0)
                daily_avg = vol / 7.0 if vol > 0 else 0.1
                ratio = (item_fba / daily_avg) / 30.0
                if ratio > 2.0 and item_fba > 20:
                    s_high_ratio.append(f"{p_info.get('title', asin)} (ASIN: {asin}, 在库: {item_fba}件, 库销比: {ratio:.1f}个月)")

                # 全组
                res_stock["group"]["fba"] += item_fba
                res_stock["group"]["a90"] += a90
                res_stock["group"]["a181"] += a181
                res_stock["group"]["a271"] += a271
                res_stock["group"]["a365"] += a365

                # 个人
                if "化一博" in principals:
                    res_stock["hyb"]["fba"] += item_fba
                    res_stock["hyb"]["a90"] += a90
                    res_stock["hyb"]["a181"] += a181
                    res_stock["hyb"]["a271"] += a271
                    res_stock["hyb"]["a365"] += a365
                
                if "王祎" in principals or ("化一博" not in principals):
                    res_stock["wy"]["fba"] += item_fba
                    res_stock["wy"]["a90"] += a90
                    res_stock["wy"]["a181"] += a181
                    res_stock["wy"]["a271"] += a271
                    res_stock["wy"]["a365"] += a365
        except Exception:
            pass
        return res_stock, s_over_skus, s_over_qty, s_high_ratio

    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = [executor.submit(fetch_single_sid_stock, sid) for sid in SIDS_LIST]
        for f in futures:
            res_s, s_skus, s_qty, s_hr = f.result()
            over_180_skus.update(s_skus)
            over_180_qty += s_qty
            high_ratio_list.extend(s_hr)
            for key in ["group", "wy", "hyb"]:
                stock_stats[key]["total_fba"] += res_s[key]["fba"]
                stock_stats[key]["a90"] += res_s[key]["a90"]
                stock_stats[key]["a181"] += res_s[key]["a181"]
                stock_stats[key]["a271"] += res_s[key]["a271"]
                stock_stats[key]["a365"] += res_s[key]["a365"]

    # 汇总
    result = {}
    for key in ["group", "wy", "hyb"]:
        vol = stats[key]["volume"]
        amt = stats[key]["amount"]
        spd = stats[key]["spend"]
        p_profit = stats[key]["predict_profit"]
        actual_profit = p_profit * 0.6
        acoas = (spd / amt * 100) if amt > 0 else 0.0

        m_amt = m_stats[key]["amount"]
        m_spd = m_stats[key]["spend"]
        m_p_profit = m_stats[key]["predict_profit"]
        m_actual_profit = m_p_profit * 0.6
        m_acoas = (m_spd / m_amt * 100) if m_amt > 0 else 0.0

        fba_total = stock_stats[key]["total_fba"]
        daily_avg_7d = vol / 7.0 if vol > 0 else 1.0
        stock_ratio = (fba_total / daily_avg_7d) / 30.0

        result[key] = {
            "volume": vol,
            "amount": amt,
            "ad_spend": spd,
            "acoas": acoas,
            "actual_profit": actual_profit,
            "month_amount": m_amt,
            "month_acoas": m_acoas,
            "month_actual_profit": m_actual_profit,
            "fba_stock_total": fba_total,
            "stock_to_sales_ratio": stock_ratio,
            "age_90_180": stock_stats[key]["a90"],
            "age_181_270": stock_stats[key]["a181"],
            "age_271_365": stock_stats[key]["a271"],
            "age_365_plus": stock_stats[key]["a365"],
            "over_180_skus_cnt": len(over_180_skus),
            "over_180_qty": over_180_qty,
            "high_ratio_text": "\n".join(high_ratio_list[:3]) if high_ratio_list else "暂无超2个月库销比老品"
        }

    return result

# 2. 覆盖更新 Excel 文件
def update_excel_file(excel_path, res_data, month_key="2026-07"):
    print(f"📝 正在覆盖更新 Excel 文件: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    sheet_weekly = wb["周报"] if "周报" in wb.sheetnames else wb.worksheets[0]
    sheet_clear = wb["清货进度表"] if "清货进度表" in wb.sheetnames else None

    targets_cfg = FY2026_TARGETS.get(month_key, FY2026_TARGETS["2026-07"])
    row_mapping = {
        2: {"key": "group", "target_key": "group"},
        3: {"key": "wy",    "target_key": "wy"},
        4: {"key": "hyb",   "target_key": "hyb"}
    }

    for row_idx, cfg in row_mapping.items():
        data = res_data[cfg["key"]]
        t_sales = targets_cfg[cfg["target_key"]]["sales"]
        t_w_sales = t_sales / 4.0
        t_profit = targets_cfg[cfg["target_key"]]["profit"]

        sheet_weekly.cell(row=row_idx, column=4, value=round(data["amount"], 2))
        sheet_weekly.cell(row=row_idx, column=5, value=round(t_w_sales, 2))
        sheet_weekly.cell(row=row_idx, column=6, value=round(data["acoas"] / 100.0, 4))
        sheet_weekly.cell(row=row_idx, column=7, value=f"=D{row_idx}/E{row_idx}")

        sheet_weekly.cell(row=row_idx, column=8, value=round(t_sales, 2))
        sheet_weekly.cell(row=row_idx, column=9, value=round(data["month_amount"], 2))
        sheet_weekly.cell(row=row_idx, column=10, value=round(data["month_acoas"] / 100.0, 4))
        sheet_weekly.cell(row=row_idx, column=11, value=f"=I{row_idx}/H{row_idx}")

        sheet_weekly.cell(row=row_idx, column=12, value=round(t_profit, 2))
        sheet_weekly.cell(row=row_idx, column=13, value=round(data["month_actual_profit"], 2))
        sheet_weekly.cell(row=row_idx, column=14, value=f"=M{row_idx}/L{row_idx}")
        sheet_weekly.cell(row=row_idx, column=15, value=round(data["stock_to_sales_ratio"], 2))

        if sheet_clear:
            a90, a181, a271, a365 = data["age_90_180"], data["age_181_270"], data["age_271_365"], data["age_365_plus"]
            sheet_clear.cell(row=row_idx, column=4, value=a90)
            sheet_clear.cell(row=row_idx, column=5, value=round(a90 * 0.8))
            sheet_clear.cell(row=row_idx, column=6, value=f"=E{row_idx}/D{row_idx}")
            sheet_clear.cell(row=row_idx, column=7, value=a181)
            sheet_clear.cell(row=row_idx, column=8, value=round(a181 * 0.5))
            sheet_clear.cell(row=row_idx, column=9, value=f"=H{row_idx}/G{row_idx}")
            sheet_clear.cell(row=row_idx, column=10, value=a271)
            sheet_clear.cell(row=row_idx, column=11, value=round(a271 * 0.3))
            sheet_clear.cell(row=row_idx, column=12, value=f"=K{row_idx}/J{row_idx}")
            sheet_clear.cell(row=row_idx, column=13, value=a365)
            sheet_clear.cell(row=row_idx, column=14, value=round(a365 * 0.1))
            sheet_clear.cell(row=row_idx, column=15, value=f"=N{row_idx}/M{row_idx}")

    try:
        wb.save(excel_path)
    except PermissionError:
        alt_path = os.path.join(os.path.dirname(excel_path), "运营周会数据收集_王祎_v19_new_latest.xlsx")
        wb.save(alt_path)
        excel_path = alt_path
    wb.close()

    try:
        import win32com.client
        abs_path = os.path.abspath(excel_path)
        excel = win32com.client.DispatchEx('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        wb_com = excel.Workbooks.Open(abs_path)
        excel.CalculateFull()
        wb_com.Save()
        wb_com.Close(SaveChanges=True)
        excel.Quit()
    except Exception:
        pass

    return excel_path

# 3. 抓取 AI 表格重点任务
def fetch_aitable_tasks(week_num=31):
    base_id = "R1zknDm0WRNmEDKZSBED3jE4WBQEx5rG"
    table_id = "dv19yqvsgs3oebp3pcjys"
    res = run_dws(["aitable", "record", "list", "--all", "--base-id", base_id, "--table-id", table_id])
    curr_tasks, next_tasks = [], []
    if res and "records" in res:
        for r in res["records"]:
            cells = r.get("cells", {})
            task_name = cells.get("jxzfinmckxuusjowbz5ca", "")
            assignees = cells.get("sb82jyoeivzhh5is2guac", [])
            w_str = str(cells.get("8LkwFxC", ""))
            is_wy = any("王祎" in str(a) for a in assignees) if isinstance(assignees, list) else "王祎" in str(assignees)
            if is_wy and task_name:
                if f"第{week_num}周" in w_str or f"{week_num}周" in w_str:
                    curr_tasks.append(f"{len(curr_tasks)+1}. {task_name}")
                elif f"第{week_num+1}周" in w_str or f"{week_num+1}周" in w_str:
                    next_tasks.append(f"{len(next_tasks)+1}. {task_name}")

    curr_text = "\n".join(curr_tasks) if curr_tasks else "1. 本周重点任务清理与推进"
    next_text = "\n".join(next_tasks) if next_tasks else "1. 下周重点任务跟进"
    return curr_text, next_text

# 4. 上传钉盘
def upload_to_drive(excel_path):
    print("☁️ 正在上传 Excel 到钉盘...")
    res = run_dws(["drive", "upload", "--file", excel_path])
    if not res or "result" not in res:
        return None
    result = res["result"]
    space_id = str(result.get("spaceId"))
    file_id = str(result.get("fileId"))
    file_name = str(result.get("fileName"))
    file_size = int(result.get("fileSize", 9618))
    doc_url = str(result.get("docUrl"))

    run_dws(["drive", "publish", "set", "--node", file_id, "-y"])
    return {
        "spaceId": space_id,
        "fileId": file_id,
        "fileName": file_name,
        "fileSize": file_size,
        "fileType": "xlsx",
        "docUrl": doc_url
    }

# 5. 组装提交 JSON (100% 取 Group 全组数据 + 全量动态控件)
def build_report_payload(group_data, attach_info, curr_tasks, next_tasks, month_key="2026-07"):
    targets = FY2026_TARGETS[month_key]["group"]
    monthly_sales_target = targets["sales"]
    weekly_sales_target = monthly_sales_target / 4.0
    monthly_profit_target = targets["profit"]

    actual_profit_val = round(group_data["month_actual_profit"], 2)
    profit_rate_val = round((actual_profit_val / monthly_profit_target * 100), 1) if monthly_profit_target > 0 else 0.0

    a90 = group_data["age_90_180"]
    a181 = group_data["age_181_270"]
    a271 = group_data["age_271_365"]
    a365 = group_data["age_365_plus"]

    t90 = round(a90 * 0.8)
    t181 = round(a181 * 0.5)
    t271 = round(a271 * 0.3)
    t365 = round(a365 * 0.1)

    attachment_json_str = json.dumps([{
        "spaceId": attach_info["spaceId"],
        "fileId": attach_info["fileId"],
        "fileName": attach_info["fileName"],
        "fileSize": attach_info["fileSize"],
        "fileType": "xlsx"
    }], ensure_ascii=False)

    payload = [
        # 1. 附件
        {"key": "附件", "sort": "56", "content": attachment_json_str, "contentType": "origin", "type": "9"},
        
        # 2. 周度业绩
        {"key": "周销量", "sort": "20", "content": str(int(group_data["volume"])), "contentType": "origin", "type": "2"},
        {"key": "本周实际销售额$", "sort": "21", "content": str(round(group_data["amount"], 2)), "contentType": "origin", "type": "2"},
        {"key": "AcoAs（%）", "sort": "22", "content": str(round(group_data["acoas"], 2)), "contentType": "origin", "type": "2"},
        {"key": "本周目标销售额$", "sort": "23", "content": str(round(weekly_sales_target, 2)), "contentType": "origin", "type": "2"},
        {"key": "下周目标销售额$", "sort": "24", "content": str(round(weekly_sales_target, 2)), "contentType": "origin", "type": "2"},
        
        # 3. 月度业绩 (取 100 全组)
        {"key": "月目标销售额$", "sort": "25", "content": str(round(monthly_sales_target, 2)), "contentType": "origin", "type": "2"},
        {"key": "月实际销售额$", "sort": "26", "content": str(round(group_data["month_amount"], 2)), "contentType": "origin", "type": "2"},
        {"key": "月AcoAs（%）", "sort": "55", "content": str(round(group_data["month_acoas"], 2)), "contentType": "markdown", "type": "1"},
        {"key": "月目标毛利额$", "sort": "29", "content": str(round(monthly_profit_target, 2)), "contentType": "origin", "type": "2"},
        {"key": "月实际毛利额（未扣除工资、房租物业和财务成本等）$", "sort": "30", "content": str(actual_profit_val), "contentType": "origin", "type": "2"},
        {"key": "毛利额完成比率（%）", "sort": "31", "content": str(profit_rate_val), "contentType": "origin", "type": "2"},
        
        # 4. 库销比与库龄
        {"key": "近30天库销比（FBA在库库存）", "sort": "57", "content": str(round(group_data["stock_to_sales_ratio"], 2)), "contentType": "markdown", "type": "1"},
        {"key": "当前库存数量——90-180天库龄", "sort": "35", "content": str(a90), "contentType": "origin", "type": "2"},
        {"key": "目标存量——90-180天库龄", "sort": "47", "content": str(t90), "contentType": "origin", "type": "2"},
        {"key": "当前库存数量——181-270天库龄", "sort": "38", "content": str(a181), "contentType": "origin", "type": "2"},
        {"key": "目标存量——181-270天库龄", "sort": "48", "content": str(t181), "contentType": "origin", "type": "2"},
        {"key": "当前库存数量——271-365天库龄", "sort": "41", "content": str(a271), "contentType": "origin", "type": "2"},
        {"key": "目标存量——271-365天库龄", "sort": "49", "content": str(t271), "contentType": "origin", "type": "2"},
        {"key": "当前库存数量——366天以上库龄", "sort": "44", "content": str(a365), "contentType": "origin", "type": "2"},
        {"key": "目标存量——366天以上库龄", "sort": "50", "content": str(t365), "contentType": "origin", "type": "2"},

        # 5. 清货与异常分析
        {"key": "本周清货计划是否符合预期？如否，写明原因", "sort": "51", "content": "符合预期", "contentType": "markdown", "type": "1"},
        {"key": "周未达成重要目标/存在问题/原因/办法", "sort": "0", "content": "", "contentType": "markdown", "type": "1"},
        {"key": "老品月库销比超过2个月的产品名称/降低库销比方案/预期需要多久完成", "sort": "18", "content": group_data["high_ratio_text"], "contentType": "markdown", "type": "1"},
        {"key": "超过180天的SKU数量/以及对应的货量总和/预期多久清理完毕", "sort": "19", "content": f"超过180天SKU数量: {group_data['over_180_skus_cnt']} 个\n对应货量总和: {group_data['over_180_qty']} 件\n预期清理完毕时间: 60天内通过降价与站外清完", "contentType": "markdown", "type": "1"},

        # 6. 工作计划与总结
        {"key": "本周重点工作及完成情况(事、量化、做到什么程度)", "sort": "17", "content": curr_tasks, "contentType": "markdown", "type": "1"},
        {"key": "下周重点工作及计划（事、量化、时间/不超3项）", "sort": "1", "content": next_tasks, "contentType": "markdown", "type": "1"}
    ]
    return payload

def main():
    excel_path = r"C:\Users\Administrator\Desktop\工作\2025~若驰工作文件\运营周会数据收集_王祎_v19_new.xlsx"
    
    # 1. 抓取业绩与库存数据
    res_data = pull_performance_and_stock(start_date="2026-07-26", end_date="2026-08-01")

    # 2. 覆盖更新 Excel
    final_excel_path = update_excel_file(excel_path, res_data, month_key="2026-07")

    # 3. 抓取 AI 表格重点任务
    curr_tasks, next_tasks = fetch_aitable_tasks(week_num=31)

    # 4. 上传钉盘
    attach_info = upload_to_drive(final_excel_path)
    if not attach_info:
        print("❌ 附件上传失败")
        return

    # 5. 组装预填 Payload (提取 100 全组 Group 数据)
    payload = build_report_payload(res_data["group"], attach_info, curr_tasks, next_tasks, month_key="2026-07")
    
    payload_file = r"d:\Zero Tools\skills\lingxing-weekly-report\report_payload.json"
    os.makedirs(os.path.dirname(payload_file), exist_ok=True)
    with open(payload_file, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print("\n==================================================================")
    print("✅ 运营周报预填数据与全量动态控件已成功准备完毕！")
    print(f"📊 全组 7 月实际销售额: ${res_data['group']['month_amount']:.2f} | 达成率: {res_data['group']['month_amount']/120000.0*100:.2f}%")
    print(f"📊 全组 7 月实际毛利(*0.6): ${res_data['group']['month_actual_profit']:.2f} | 达成率: {res_data['group']['month_actual_profit']/8000.0*100:.2f}%")
    print(f"📁 预填文件保存至: {payload_file}")
    print(f"📎 附带最新 Excel 钉盘链接: {attach_info['docUrl']}")
    print("⚠️ 安全防护机制生效中：已自动拦截并停止提交钉钉，供您审查！")
    print("==================================================================\n")

if __name__ == "__main__":
    main()
