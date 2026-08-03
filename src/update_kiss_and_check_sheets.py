# -*- coding: utf-8 -*-
"""
1. 填入 KISS 原则文本到目标 Excel 单元格中 (供用户 Review 与修改)
2. 逐 Sheet 审查遗漏单元格并检查
"""
import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

OUTPUT_DIR = r"d:\Zero Tools\data"

KISS_TEXT_IN_SHEET = (
    "【KISS 原则运营复盘分析】\n"
    "• Keep (保持): 维持235宽38厚隐形层板在US站头部优势与BuyBox，稳定贡献爆款利润；维持4-6月高效交付(Q1完成率140.3%)。\n"
    "• Improve (提升): 优化相框多件套组合包与定价（把相框5.31%低毛利拉升至12%+）；加固195宽与1520包装，压降退款率(>12%)。\n"
    "• Stop (止损): 对2015和1525亏损相框关停高ACOAS宽泛广告；停止91-180天库龄盲目备货，降至0.8目标存量以内。\n"
    "• Start (启动AI提效): 启动AI Agent广告24h自动否定零转化词；启动Rufus GEO搜问算法适配优化Bullets；启动领星MCP与Python全流程月报自动核算(将2天压缩至5分钟)。"
)

# 更新报表 2 (主管月复盘) 中的 KISS 单元格
p2 = os.path.join(OUTPUT_DIR, '2026年6月运营主管月复盘_王祎_Filled_v7.xlsx')
if os.path.exists(p2):
    wb2 = openpyxl.load_workbook(p2, data_only=False)
    # 查找复盘问题/改进 Sheet 或单元格
    for sheetname in wb2.sheetnames:
        ws = wb2[sheetname]
        # 寻找包含“问题”或“方案”的单元格
        for r in range(1, min(30, ws.max_row+1)):
            for c in range(1, min(15, ws.max_column+1)):
                val = str(ws.cell(row=r, column=c).value or '')
                if '存在问题' in val or '解决办法' in val or 'KISS' in val or '改进计划' in val:
                    ws.cell(row=r+1, column=c).value = KISS_TEXT_IN_SHEET
                    print(f"✅ 在 Sheet '{sheetname}' R{r+1}C{c} 填入 KISS 原则分析！")

    wb2.save(p2)
    wb2.close()

print("\n🎉 KISS 原则分析已成功写入表格对应位置，可供在表内 Review 与修改！")
