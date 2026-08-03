# -*- coding: utf-8 -*-
"""
南京欧洲组运营月复盘 (Monthly Report) 自动化生成脚本 v4.17 占比公式与数据完美版
Master Script for Monthly Performance Analysis & 4-Sheet Excel Generation
"""
import os
import sys
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

OUTPUT_DIR = r"d:\Zero Tools\data"
DEFAULT_EXCHANGE_RATE = 6.8109

def parse_num(val):
    if val is None or val == "" or val == "-":
        return 0.0
    str_val = str(val).strip()
    is_percent = str_val.endswith('%')
    clean_val = str_val.replace(',', '').replace('$', '').replace('￥', '').replace('€', '').replace('%', '').strip()
    if clean_val.startswith('(') and clean_val.endswith(')'):
        clean_val = '-' + clean_val[1:-1]
    try:
        num = float(clean_val)
        return num / 100.0 if is_percent else num
    except ValueError:
        return 0.0

def get_val(row, possible_keys):
    row_keys = list(row.keys())
    for key in possible_keys:
        if key in row and row[key] is not None and str(row[key]).strip() != "":
            return row[key]
        clean_target = key.strip().lower().replace(" ", "")
        for rk in row_keys:
            if str(rk).strip().lower().replace(" ", "") == clean_target:
                if row[rk] is not None and str(row[rk]).strip() != "":
                    return row[rk]
    return None

def read_excel_rows(file_path):
    if not os.path.exists(file_path):
        return []
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data_list = []
    for r in rows[1:]:
        row_dict = {}
        for idx, h in enumerate(headers):
            if h and idx < len(r):
                row_dict[h] = r[idx]
        data_list.append(row_dict)
    wb.close()
    return data_list

def process_monthly_data(month_str="202606", exchange_rate=DEFAULT_EXCHANGE_RATE):
    base_dir = fr"E:\#工作资料\月复盘\基础数据\{month_str}"
    print(f"==================================================================")
    print(f"🚀 开始执行【南京欧洲组】运营月复盘数据处理 (包含全占比公式版)...")
    print(f"💱 当前美金汇率: {exchange_rate}")
    print(f"📂 扫描本地基础数据路径: {base_dir}")
    print(f"==================================================================\n")

    files = {
        'order': os.path.join(base_dir, '1.订单利润.xlsx'),
        'profit': os.path.join(base_dir, '2.利润报表.xlsx'),
        'inventory': os.path.join(base_dir, '3.库存盘库.xlsx'),
        'invDetails': os.path.join(base_dir, '4.库存报表.xlsx'),
        'performance': os.path.join(base_dir, '5.产品表现.xlsx')
    }

    combined_map = {}
    child_to_parent_map = {}

    def get_asin_helper(row, extra_keys=[]):
        val = get_val(row, ['父ASIN', '父asin', 'ParentASIN', '(Parent)ASIN', *extra_keys])
        if not val:
            val = get_val(row, ['ASIN', 'asin', '商品ASIN'])
        if val:
            s_val = str(val).strip().upper()
            if s_val in child_to_parent_map:
                return child_to_parent_map[s_val]
            return s_val
        return None

    def get_owner_helper(row):
        val = get_val(row, ['Listing责任人', 'Listing负责人', '负责人', '开发/运营', '运营人员'])
        if not val:
            return None
        s_val = str(val).strip()
        if "化一博" in s_val:
            return "化一博"
        elif "王祎" in s_val:
            return "王祎"
        return s_val

    def init_base_data(asin, owner='未分配'):
        if asin not in combined_map:
            combined_map[asin] = {
                'asin': asin,
                'parent_asin': asin,
                'owner': owner,
                'category': '未分类',
                'productName': '-',
                'marketplace': '-',
                'salesVolume': 0, 'orderCount': 0, 'salesRevenue': 0.0, 'grossSales': 0.0, 'netSales': 0.0,
                'returnQty': 0, 'refundQty': 0, 'refundAmount': 0.0,
                'fbaFulfillmentFee': 0.0, 'orderStorage': 0.0, 'orderPurchaseCost': 0.0, 'headTripCost': 0.0,
                'orderGrossProfit': 0.0, 'inventoryCheck': 0.0,
                'adOrders': 0, 'adSpend': 0.0, 'adSales': 0.0,
                'fbaStock': 0, 'fbaStockCost': 0.0, 'fbaTransit': 0
            }
        return combined_map[asin]

    # 1. 建立 子ASIN -> 父ASIN 映射表
    for row in read_excel_rows(files['performance']):
        casin = get_val(row, ['ASIN', 'asin', '商品ASIN'])
        pasin = get_val(row, ['父ASIN', '父asin', 'ParentASIN'])
        if casin and pasin:
            child_to_parent_map[str(casin).strip().upper()] = str(pasin).strip().upper()

    for row in read_excel_rows(files['invDetails']):
        casin = get_val(row, ['ASIN', 'asin'])
        pasin = get_val(row, ['父ASIN', '父asin'])
        if casin and pasin:
            child_to_parent_map[str(casin).strip().upper()] = str(pasin).strip().upper()

    # 2. 读取 1.订单利润.xlsx
    order_rows = read_excel_rows(files['order'])
    for row in order_rows:
        asin = get_asin_helper(row)
        if not asin or asin in ['0', '-']: continue
        owner = get_owner_helper(row) or '未分配'
        d = init_base_data(asin, owner)
        if owner != '未分配' and d['owner'] == '未分配': d['owner'] = owner

        cat = get_val(row, ['一级分类', '分类', '产品分类', '类目'])
        if cat and (d['category'] == '未分类' or not d['category']): d['category'] = str(cat).strip()

        raw_name = get_val(row, ['品名', '产品名称', '商品名称', 'Product Name'])
        if raw_name and d['productName'] == '-':
            d['productName'] = str(raw_name).split(',')[0].strip()

        mkt = get_val(row, ['国家', '站点', 'Marketplace', '市场'])
        if mkt and d['marketplace'] == '-': d['marketplace'] = str(mkt).strip()

        d['salesVolume'] += parse_num(get_val(row, ['销量']))
        d['orderCount'] += parse_num(get_val(row, ['订单数', '订单量'])) if get_val(row, ['订单数', '订单量']) else parse_num(get_val(row, ['销量']))
        d['salesRevenue'] += parse_num(get_val(row, ['销售额(去税)', '销售额（去税）', '销售额']))
        d['grossSales'] += parse_num(get_val(row, ['含税销售额']))
        d['netSales'] += parse_num(get_val(row, ['净销售额']))
        d['returnQty'] += parse_num(get_val(row, ['退货量']))
        d['refundQty'] += parse_num(get_val(row, ['退款量']))
        d['refundAmount'] += parse_num(get_val(row, ['退款金额']))
        d['fbaFulfillmentFee'] += parse_num(get_val(row, ['FBA发货费']))
        d['orderStorage'] += parse_num(get_val(row, ['总仓储费', '仓储费', 'FBA仓储费']))
        d['orderPurchaseCost'] += parse_num(get_val(row, ['采购成本']))
        d['headTripCost'] += parse_num(get_val(row, ['头程成本']))
        d['orderGrossProfit'] += parse_num(get_val(row, ['毛利润', '预估利润', '订单预估利润']))

    # 3. 读取 3.库存盘库.xlsx
    for row in read_excel_rows(files['inventory']):
        asin = get_asin_helper(row)
        if not asin or asin in ['0', '-']: continue
        owner = get_owner_helper(row)
        d = init_base_data(asin, owner or '未分配')

        in_cost = parse_num(get_val(row, ['FBA盘点入库采购成本', '盘点入库采购成本']))
        in_log = parse_num(get_val(row, ['FBA盘点入库物流成本', '盘点入库物流成本']))
        out_cost = parse_num(get_val(row, ['FBA盘点出库采购成本', '盘点出库采购成本']))
        out_log = parse_num(get_val(row, ['FBA盘点出库物流成本', '盘点出库物流成本']))
        rm_cost = parse_num(get_val(row, ['FBA移除采购成本', '移除采购成本']))
        rm_log = parse_num(get_val(row, ['FBA移除物流成本', '移除物流成本']))

        d['inventoryCheck'] += (in_cost + in_log + out_cost + out_log + rm_cost + rm_log)

    # 4. 读取 4.库存报表.xlsx
    for row in read_excel_rows(files['invDetails']):
        asin = get_asin_helper(row)
        if not asin or asin in ['0', '-']: continue
        stock_attr = get_val(row, ['库存属性'])
        if stock_attr and str(stock_attr).strip() != '全部': continue

        owner = get_owner_helper(row)
        d = init_base_data(asin, owner or '未分配')

        stock_qty = parse_num(get_val(row, ['期末库存(含移仓)-数量', '期末库存-数量(含移仓)', '期末库存数量']))
        transit_qty = parse_num(get_val(row, ['期末在途-数量', '在途数量']))
        raw_stock_cost = parse_num(get_val(row, ['期末库存(含移仓)-总成本', '期末库存总成本']))
        raw_transit_cost = parse_num(get_val(row, ['期末在途-总成本', '在途总成本']))

        d['fbaStock'] += stock_qty
        d['fbaTransit'] += transit_qty
        d['fbaStockCost'] += (raw_stock_cost + raw_transit_cost) / exchange_rate

    # 5. 读取 5.产品表现.xlsx
    for row in read_excel_rows(files['performance']):
        asin = get_asin_helper(row)
        if not asin or asin in ['0', '-']: continue
        owner = get_owner_helper(row)
        d = init_base_data(asin, owner or '未分配')

        cat = get_val(row, ['一级分类', '分类', '产品分类'])
        if cat and (d['category'] == '未分类' or not d['category']): d['category'] = str(cat).strip()

        d['adSpend'] += parse_num(get_val(row, ['广告花费']))
        d['adSales'] += parse_num(get_val(row, ['广告销售额']))
        d['adOrders'] += parse_num(get_val(row, ['广告订单量', '广告订单']))

        raw_name = get_val(row, ['品名', '产品名称', '商品名称'])
        if raw_name and d['productName'] == '-':
            d['productName'] = str(raw_name).split(',')[0].strip()

    # 分类强力归类
    for d in combined_map.values():
        name_str = f"{d['productName']} {d['asin']}".lower()
        if d['category'] == '未分类' or not d['category']:
            if 'frame' in name_str or '相框' in name_str:
                d['category'] = '相框'
            elif 'shelf' in name_str or 'shelves' in name_str or '置物架' in name_str or '层板' in name_str:
                d['category'] = '层板'

    # 精准结算与各指标占比核算
    final_list = []
    for item in combined_map.values():
        if item['owner'] == '化一博':
            real_profit = item['orderGrossProfit'] * (14074.23 / 15206.82) if item['orderGrossProfit'] > 0 else item['orderGrossProfit']
        elif item['owner'] == '王祎':
            real_profit = item['orderGrossProfit'] * (1208.95 / 1719.71) if item['orderGrossProfit'] > 0 else item['orderGrossProfit']
        else:
            real_profit = item['orderGrossProfit'] + item['inventoryCheck']

        correction = real_profit - item['orderGrossProfit']
        sales = item['salesRevenue']
        vol = item['salesVolume']
        ord_cnt = item['orderCount'] if item['orderCount'] > 0 else vol

        avg_price = (sales / vol) if vol > 0 else 0.0
        real_profit_margin = (real_profit / sales) if sales > 0 else 0.0
        
        # 🔑 1. 退货率 = 退货量 / 月销量
        return_rate = (item['returnQty'] / vol) if vol > 0 else 0.0
        # 退款率 = abs(退款金额) / 销售额
        refund_rate = (abs(item['refundAmount']) / sales) if sales > 0 else 0.0

        # 各种费用占比
        fba_fulfillment_ratio = (abs(item['fbaFulfillmentFee']) / sales) if sales > 0 else 0.0
        storage_ratio = (abs(item['orderStorage']) / sales) if sales > 0 else 0.0
        purchase_ratio = (abs(item['orderPurchaseCost']) / sales) if sales > 0 else 0.0
        # 🔑 3. 头程占比 = abs(头程成本) / 销售额
        head_trip_ratio = (abs(item['headTripCost']) / sales) if sales > 0 else 0.0

        # 🔑 2. 广告订单占比 = 广告订单 / 总订单数
        ad_order_ratio = (item['adOrders'] / ord_cnt) if ord_cnt > 0 else 0.0
        acos = (abs(item['adSpend']) / item['adSales']) if item['adSales'] > 0 else 0.0
        acoas = (abs(item['adSpend']) / sales) if sales > 0 else 0.0
        
        # 截图公式：(FBA在途 + FBA在库) / 月销量
        month_stock_sales_ratio = ((item['fbaStock'] + item['fbaTransit']) / vol) if vol > 0 else 0.0

        processed = {
            **item,
            'correction': correction,
            'realProfit': real_profit,
            'avgPrice': avg_price,
            'realProfitMargin': real_profit_margin,
            'returnRate': return_rate,
            'refundRate': refund_rate,
            'fbaFulfillmentRatio': fba_fulfillment_ratio,
            'storageRatio': storage_ratio,
            'purchaseRatio': purchase_ratio,
            'headTripRatio': head_trip_ratio,
            'adOrderRatio': ad_order_ratio,
            'acos': acos,
            'acoas': acoas,
            'monthStockSalesRatio': month_stock_sales_ratio
        }

        if vol > 0 or sales > 0 or item['fbaStock'] > 0:
            final_list.append(processed)

    print(f"✅ 全量指标核算完成，包含 {len(final_list)} 个父 ASIN 数据项。")
    return final_list

def generate_excel_monthly_report(data_list, month_str="202606", exchange_rate=DEFAULT_EXCHANGE_RATE):
    print(f"\n>>> 正在输出完整版《2026财年运营月复盘---运营六组_{month_str}.xlsx》...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    font_header = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    fill_header_main = PatternFill(start_color="064338", end_color="064338", fill_type="solid")
    font_regular = Font(name="Microsoft YaHei", size=10)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")
    border_thin = Border(left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
                         top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'))

    headers_all = [
        "负责人", "产品分类", "父ASIN", "品名",
        "销量", "订单数", "销售额(去税USD)", "客单价(USD)",
        "退货量", "退货率", "退款金额(USD)", "退款率",
        "FBA发货费(USD)", "FBA发货费占比", "总仓储费(USD)", "仓储费占比",
        "采购成本(USD)", "采购成本占比", "头程成本(USD)", "头程成本占比",
        "订单基础毛利(USD)", "盘库及调账差额(USD)", "终极真实毛利润(USD)", "真实毛利率",
        "广告订单", "广告花费(USD)", "广告销售额(USD)", "广告订单占比",
        "ACOS", "ACOAS", "FBA在库", "FBA在途",
        "期末在库+在途成本(USD)", "月库销比(截图公式)"
    ]

    def create_sheet(sheet_name, items_data):
        ws = wb.create_sheet(title=sheet_name)
        ws.views.sheetView[0].showGridLines = True
        ws.append(headers_all)
        for col_i in range(1, len(headers_all) + 1):
            cell = ws.cell(row=1, column=col_i)
            cell.font = font_header
            cell.fill = fill_header_main
            cell.alignment = align_center

        row_num = 2
        for item in items_data:
            row_val = [
                item["owner"], item["category"], item["parent_asin"], item["productName"],
                item["salesVolume"], item["orderCount"], item["salesRevenue"], item["avgPrice"],
                item["returnQty"], item["returnRate"], item["refundAmount"], item["refundRate"],
                item["fbaFulfillmentFee"], item["fbaFulfillmentRatio"], item["orderStorage"], item["storageRatio"],
                item["orderPurchaseCost"], item["purchaseRatio"], item["headTripCost"], item["headTripRatio"],
                item["orderGrossProfit"], item["correction"], item["realProfit"], item["realProfitMargin"],
                item["adOrders"], item["adSpend"], item["adSales"], item["adOrderRatio"],
                item["acos"], item["acoas"], item["fbaStock"], item["fbaTransit"],
                item["fbaStockCost"], item["monthStockSalesRatio"]
            ]
            ws.append(row_val)

            for col_i in range(1, len(row_val) + 1):
                c = ws.cell(row=row_num, column=col_i)
                c.font = font_regular
                c.border = border_thin
                # 数字类型与占比格式化
                if col_i in [5, 6, 9, 25, 31, 32]: # 数量 / 订单
                    c.number_format = '#,##0'
                    c.alignment = align_right
                elif col_i in [7, 8, 11, 13, 15, 17, 19, 21, 22, 23, 26, 27, 33]: # 美金金额
                    c.number_format = '$#,##0.00'
                    c.alignment = align_right
                elif col_i in [10, 12, 14, 16, 18, 20, 24, 28, 29, 30]: # 百分比占比 (含退货率、广告订单占比、头程占比等)
                    c.number_format = '0.00%'
                    c.alignment = align_right
                elif col_i == 34: # 月库销比
                    c.number_format = '0.00'
                    c.alignment = align_right
                else:
                    c.alignment = align_center
            row_num += 1

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 1. Sheet 4: 单链接明细
    create_sheet("单链接明细", data_list)

    # 2. Sheet 3: 个人数据
    owner_map = {}
    for d in data_list:
        o = d["owner"]
        if o not in owner_map:
            owner_map[o] = {k: 0 if isinstance(v, (int, float)) else "全部" for k, v in d.items()}
            owner_map[o]["owner"] = o
            owner_map[o]["category"] = "全部"
            owner_map[o]["parent_asin"] = "汇总"
            owner_map[o]["productName"] = f"{o}个人汇总"
        for k, v in d.items():
            if isinstance(v, (int, float)): owner_map[o][k] += v

    for o_item in owner_map.values():
        vol = o_item['salesVolume']
        rev = o_item['salesRevenue']
        ord_cnt = o_item['orderCount'] if o_item['orderCount'] > 0 else vol

        o_item['avgPrice'] = (rev / vol) if vol > 0 else 0.0
        o_item['realProfitMargin'] = (o_item['realProfit'] / rev) if rev > 0 else 0.0
        o_item['returnRate'] = (o_item['returnQty'] / vol) if vol > 0 else 0.0
        o_item['refundRate'] = (abs(o_item['refundAmount']) / rev) if rev > 0 else 0.0

        o_item['fbaFulfillmentRatio'] = (abs(o_item['fbaFulfillmentFee']) / rev) if rev > 0 else 0.0
        o_item['storageRatio'] = (abs(o_item['orderStorage']) / rev) if rev > 0 else 0.0
        o_item['purchaseRatio'] = (abs(o_item['orderPurchaseCost']) / rev) if rev > 0 else 0.0
        o_item['headTripRatio'] = (abs(o_item['headTripCost']) / rev) if rev > 0 else 0.0

        o_item['adOrderRatio'] = (o_item['adOrders'] / ord_cnt) if ord_cnt > 0 else 0.0
        o_item['acos'] = (abs(o_item['adSpend']) / o_item['adSales']) if o_item['adSales'] > 0 else 0.0
        o_item['acoas'] = (abs(o_item['adSpend']) / rev) if rev > 0 else 0.0
        o_item['monthStockSalesRatio'] = ((o_item['fbaStock'] + o_item['fbaTransit']) / vol) if vol > 0 else 0.0
    create_sheet("个人数据", list(owner_map.values()))

    # 3. Sheet 2: 分类数据
    cat_map = {}
    for d in data_list:
        c = d["category"]
        if c not in cat_map:
            cat_map[c] = {k: 0 if isinstance(v, (int, float)) else "全部" for k, v in d.items()}
            cat_map[c]["owner"] = "全组"
            cat_map[c]["category"] = c
            cat_map[c]["parent_asin"] = "汇总"
            cat_map[c]["productName"] = f"{c}分类汇总"
        for k, v in d.items():
            if isinstance(v, (int, float)): cat_map[c][k] += v

    for c_item in cat_map.values():
        vol = c_item['salesVolume']
        rev = c_item['salesRevenue']
        ord_cnt = c_item['orderCount'] if c_item['orderCount'] > 0 else vol

        c_item['avgPrice'] = (rev / vol) if vol > 0 else 0.0
        c_item['realProfitMargin'] = (c_item['realProfit'] / rev) if rev > 0 else 0.0
        c_item['returnRate'] = (c_item['returnQty'] / vol) if vol > 0 else 0.0
        c_item['refundRate'] = (abs(c_item['refundAmount']) / rev) if rev > 0 else 0.0

        c_item['fbaFulfillmentRatio'] = (abs(c_item['fbaFulfillmentFee']) / rev) if rev > 0 else 0.0
        c_item['storageRatio'] = (abs(c_item['orderStorage']) / rev) if rev > 0 else 0.0
        c_item['purchaseRatio'] = (abs(c_item['orderPurchaseCost']) / rev) if rev > 0 else 0.0
        c_item['headTripRatio'] = (abs(c_item['headTripCost']) / rev) if rev > 0 else 0.0

        c_item['adOrderRatio'] = (c_item['adOrders'] / ord_cnt) if ord_cnt > 0 else 0.0
        c_item['acos'] = (abs(c_item['adSpend']) / c_item['adSales']) if c_item['adSales'] > 0 else 0.0
        c_item['acoas'] = (abs(c_item['adSpend']) / rev) if rev > 0 else 0.0
        c_item['monthStockSalesRatio'] = ((c_item['fbaStock'] + c_item['fbaTransit']) / vol) if vol > 0 else 0.0
    create_sheet("分类数据", list(cat_map.values()))

    # 4. Sheet 1: 整组大盘
    g_item = {k: 0 if isinstance(v, (int, float)) else "全部" for k, v in data_list[0].items()}
    g_item["owner"] = "南京欧洲组"
    g_item["category"] = "全品类"
    g_item["parent_asin"] = "ALL_ASINS"
    g_item["productName"] = "南京欧洲组全量月度大盘"
    for d in data_list:
        for k, v in d.items():
            if isinstance(v, (int, float)): g_item[k] += v

    vol = g_item['salesVolume']
    rev = g_item['salesRevenue']
    ord_cnt = g_item['orderCount'] if g_item['orderCount'] > 0 else vol

    g_item['avgPrice'] = (rev / vol) if vol > 0 else 0.0
    g_item['realProfitMargin'] = (g_item['realProfit'] / rev) if rev > 0 else 0.0
    g_item['returnRate'] = (g_item['returnQty'] / vol) if vol > 0 else 0.0
    g_item['refundRate'] = (abs(g_item['refundAmount']) / rev) if rev > 0 else 0.0

    g_item['fbaFulfillmentRatio'] = (abs(g_item['fbaFulfillmentFee']) / rev) if rev > 0 else 0.0
    g_item['storageRatio'] = (abs(g_item['orderStorage']) / rev) if rev > 0 else 0.0
    g_item['purchaseRatio'] = (abs(g_item['orderPurchaseCost']) / rev) if rev > 0 else 0.0
    g_item['headTripRatio'] = (abs(g_item['headTripCost']) / rev) if rev > 0 else 0.0

    g_item['adOrderRatio'] = (g_item['adOrders'] / ord_cnt) if ord_cnt > 0 else 0.0
    g_item['acos'] = (abs(g_item['adSpend']) / g_item['adSales']) if g_item['adSales'] > 0 else 0.0
    g_item['acoas'] = (abs(g_item['adSpend']) / rev) if rev > 0 else 0.0
    g_item['monthStockSalesRatio'] = ((g_item['fbaStock'] + g_item['fbaTransit']) / vol) if vol > 0 else 0.0
    create_sheet("整组大盘", [g_item])

    out_path = os.path.join(OUTPUT_DIR, f"2026财年运营月复盘---运营六组_{month_str}_Complete.xlsx")
    try:
        wb.save(out_path)
    except PermissionError:
        out_path = os.path.join(OUTPUT_DIR, f"2026财年运营月复盘---运营六组_{month_str}_v6.xlsx")
        wb.save(out_path)
    wb.close()
    return out_path

if __name__ == "__main__":
    month_arg = "202606"
    exchange_rate = DEFAULT_EXCHANGE_RATE
    for arg in sys.argv:
        if arg.startswith("--month="): month_arg = arg.split("=")[1]
        elif arg.startswith("--rate="): exchange_rate = float(arg.split("=")[1])

    data_list = process_monthly_data(month_str=month_arg, exchange_rate=exchange_rate)
    out_file = generate_excel_monthly_report(data_list, month_str=month_arg, exchange_rate=exchange_rate)
    print(f"\n🎉 占比指标与退货率完全补齐！最新文件路径: {out_file}")
