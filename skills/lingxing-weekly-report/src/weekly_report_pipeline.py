# -*- coding: utf-8 -*-
"""
KovaScape 领星周报统一流水线 (Master Weekly Report Pipeline)
支持:
  - 自动日期推算 (基于“当前系统时间”或可选 week-end 指定周六)
  - 报表一: fill_weekly_report（运营周复盘，更新 Excel + 上传钉盘 + 生成 report_payload.json 挂载附件）
  - 报表二: generate_weekly_meeting_report（周会会议纪要及数据环比在线表 + 自动推送钉群消息）
  - 支持 --dry-run
"""
import os
import sys
import json
import threading
import datetime
import urllib.request
import openpyxl
import subprocess
from concurrent.futures import ThreadPoolExecutor

# Ensure UTF-8 output
sys.stdout.reconfigure(encoding='utf-8')

# 常量定义
DWS_BIN = r"D:\Zero Tools\DingTalk\bin\dws.exe"
MCP_CONFIG_PATH = r"C:\Users\Administrator\.gemini\config\mcp_config.json"
MCP_URL = "https://openmcp.lingxing.com/mcp-servers/lingxing-mcp"
REPORT_TEMPLATE_ID = "17a14a44cdee2e409b88ad14ca68d77b" # 运营周复盘（周一9:00前提交）
CHAT_ID = "cidCtsmbs4Sk6ajOZQDMQl32w==" # 南京欧洲站￥$€£
USER_ID_WY = "17566881508928543"
USER_ID_HYB = "17597998065506150"

SIDS_STR = "5030,5751,5019,5024,5026,5025,5031,5023,5021,5020,5027,5029,5028,5022,5018"
SIDS_LIST = [5030, 5751, 5019, 5024, 5026, 5025, 5031, 5023, 5021, 5020, 5027, 5029, 5028, 5022, 5018]

BASE_ID = "R1zknDm0WRNmEDKZSBED3jE4WBQEx5rG" # 周重点任务
TABLE_ID = "dv19yqvsgs3oebp3pcjys"          # 1.任务管理表

# Excel 模板及输出路径
EXCEL_TEMPLATE_R1 = r"E:\#工作资料\月复盘\运营周会数据收集_王祎.xlsx"
EXCEL_TEMPLATE_R2 = r"D:\Zero Tools\data\六组周会会议纪要20260727.xlsx"
OUTPUT_DIR = r"d:\Zero Tools\data"

# FY2026 财年内置目标表 (包含销售额与毛利额)
FY2026_TARGETS = {
    "2026-04": {
        "group": {"sales": 150000, "profit": 12000},
        "wy":    {"sales": 45000,  "profit": 3600},
        "hyb":   {"sales": 105000, "profit": 8400}
    },
    "2026-05": {
        "group": {"sales": 90000, "profit": 6000},
        "wy":    {"sales": 27000,  "profit": 1800},
        "hyb":   {"sales": 63000,  "profit": 4200}
    },
    "2026-06": {
        "group": {"sales": 65000, "profit": 5000},
        "wy":    {"sales": 19500,  "profit": 1500},
        "hyb":   {"sales": 45500,  "profit": 3500}
    },
    "2026-07": {
        "group": {"sales": 120000, "profit": 8000},
        "wy":    {"sales": 48000,  "profit": 3600},
        "hyb":   {"sales": 72000,  "profit": 6400}
    },
    "2026-08": {
        "group": {"sales": 280000, "profit": 40000},
        "wy":    {"sales": 112000, "profit": 16000},
        "hyb":   {"sales": 168000, "profit": 24000}
    },
    "2026-09": {
        "group": {"sales": 280000, "profit": 40000},
        "wy":    {"sales": 112000, "profit": 16000},
        "hyb":   {"sales": 168000, "profit": 24000}
    },
    "2026-10": {
        "group": {"sales": 450000, "profit": 36000},
        "wy":    {"sales": 180000, "profit": 14400},
        "hyb":   {"sales": 270000, "profit": 21600}
    },
    "2026-11": {
        "group": {"sales": 550000, "profit": 60000},
        "wy":    {"sales": 220000, "profit": 24000},
        "hyb":   {"sales": 330000, "profit": 36000}
    },
    "2026-12": {
        "group": {"sales": 600000, "profit": 70000},
        "wy":    {"sales": 270000, "profit": 31500},
        "hyb":   {"sales": 330000, "profit": 38500}
    },
    "2027-01": {
        "group": {"sales": 720000, "profit": 100000},
        "wy":    {"sales": 324000, "profit": 45000},
        "hyb":   {"sales": 396000, "profit": 55000}
    },
    "2027-02": {
        "group": {"sales": 880000, "profit": 115000},
        "wy":    {"sales": 396000, "profit": 51750},
        "hyb":   {"sales": 484000, "profit": 63250}
    },
    "2027-03": {
        "group": {"sales": 950000, "profit": 126000},
        "wy":    {"sales": 427500, "profit": 56700},
        "hyb":   {"sales": 522500, "profit": 69300}
    }
}

def get_mcp_headers():
    key = "a12e733b81a539ab08cb05f0110c5624"
    if os.path.exists(MCP_CONFIG_PATH):
        try:
            with open(MCP_CONFIG_PATH, "r", encoding="utf-8") as f:
                cfg = json.load(f)
                key = cfg.get("mcpServers", {}).get("LingXing-MCP", {}).get("headers", {}).get("X-Mcp-Key", key)
        except Exception:
            pass
    return {
        "Content-Type": "application/json",
        "X-Mcp-Key": key
    }

def call_mcp_tool(name, arguments):
    payload = {
        "jsonrpc": "2.0",
        "method": "tools/call",
        "params": {
            "name": name,
            "arguments": arguments
        },
        "id": 1
    }
    req = urllib.request.Request(MCP_URL, data=json.dumps(payload).encode('utf-8'), headers=get_mcp_headers())
    with urllib.request.urlopen(req, timeout=40) as resp:
        res_json = json.loads(resp.read().decode('utf-8'))
        content_text = res_json["result"]["content"][0]["text"]
        return json.loads(content_text)

def run_dws(args):
    cmd = [DWS_BIN] + args + ["-f", "json"]
    res = subprocess.run(cmd, capture_output=True, text=True, encoding='utf-8')
    if res.returncode != 0:
        print(f"dws error: {res.stderr}")
        return None
    try:
        return json.loads(res.stdout)
    except Exception as e:
        print(f"JSON decode error: {e}")
        return None

# ========== 在线表格导入与授权（2026-08-09 修正） ==========
# 背景：钉盘文件（drive upload）是二进制附件，无法设置 EDITOR 角色权限
# （drive permission 仅适用于文档空间节点，报 "can't set role"）。
# 正确做法：doc import 导入为在线电子表格 → doc +access-grant 授予编辑权限。
DOC_IMPORT_WORKSPACE = "27227639280"   # 我的文件（当前账号空间，已验证可导入）
EDITOR_NAMES = "王祎,化一博,高应婷"      # 授予编辑权限的成员（姓名，dws 自动解析；高应婷 2026-08-09 用户要求）
# dws.js（cli-connector-packages）作为授权回退通道：
# dws.exe 需 ≥v1.0.57（2026-08-06）才支持 doc +access-grant，旧版会报 unknown flag
DWS_BIN_JS = r"C:\Users\china\.workbuddy\binaries\node\cli-connector-packages\node_modules\dingtalk-workspace-cli\bin\dws.js"

def run_dws_js(args):
    """用新版 dws.js 执行命令（支持 +access-grant 等增强命令）。"""
    cmd = ["node", DWS_BIN_JS] + args + ["-f", "json"]
    res = subprocess.run(cmd, capture_output=True, text=True, encoding='utf-8')
    if res.returncode != 0:
        print(f"dws.js error: {res.stderr}")
        return None
    try:
        return json.loads(res.stdout)
    except Exception as e:
        print(f"dws.js JSON decode error: {e}")
        return None

def grant_editor(doc_url):
    """授予 doc_url 编辑权限（EDITOR）。优先 dws.exe（≥v1.0.57 支持），失败回退 dws.js。"""
    grant = run_dws(["doc", "+access-grant", "--node", doc_url, "--to", EDITOR_NAMES, "--role", "EDITOR", "-y"])
    if grant and grant.get("status") == "success":
        return grant
    # 回退：dws.exe 旧版不支持 +access-grant 时用 dws.js
    print("⚠️ dws.exe 授权未成功，回退 dws.js ...")
    grant = run_dws_js(["doc", "+access-grant", "--node", doc_url, "--to", EDITOR_NAMES, "--role", "EDITOR", "-y"])
    return grant

def import_excel_as_online(file_path, display_name=None):
    """导入本地 Excel 为钉钉在线电子表格，并授予王祎/化一博 EDITOR 权限。
    返回 (doc_url, document_type) 或 (None, None)。"""
    args = ["doc", "import", "--file", file_path, "--workspace", DOC_IMPORT_WORKSPACE]
    if display_name:
        args += ["--name", display_name]
    res = run_dws(args)
    if not res or not res.get("success"):
        print(f"⚠️ doc import 失败: {res}")
        return None, None
    doc_url = res.get("documentUrl")
    if doc_url:
        grant = grant_editor(doc_url)
        if grant and grant.get("status") == "success":
            print(f"✅ 在线表格已授权可编辑: {EDITOR_NAMES} → {doc_url}")
        else:
            print(f"⚠️ 授权失败: {grant}")
    return doc_url, res.get("documentType")

# 日期推算逻辑
def calculate_dates(custom_week_end=None, custom_meeting_date=None):
    if custom_week_end:
        week_end = datetime.datetime.strptime(custom_week_end, "%Y-%m-%d").date()
    else:
        # 基于今天寻找最近的周六
        today = datetime.date.today()
        # today.weekday() returns 0 for Monday to 6 for Sunday
        # Saturday is 5
        offset = (today.weekday() - 5) % 7
        week_end = today - datetime.timedelta(days=offset)

    week_start = week_end - datetime.timedelta(days=6)
    month_start = week_end.replace(day=1)
    month_end = week_end
    month_key = month_end.strftime("%Y-%m")

    if custom_meeting_date:
        meeting_date = datetime.datetime.strptime(custom_meeting_date, "%Y-%m-%d").date()
    else:
        meeting_date = week_end + datetime.timedelta(days=2) # 下周一

    return {
        "week_start": week_start.strftime("%Y-%m-%d"),
        "week_end": week_end.strftime("%Y-%m-%d"),
        "month_start": month_start.strftime("%Y-%m-%d"),
        "month_end": month_end.strftime("%Y-%m-%d"),
        "month_key": month_key,
        "meeting_date": meeting_date.strftime("%Y-%m-%d")
    }

def fetch_perf_all(sids, start_date, end_date):
    """分页拉全 query_product_performance_asin_lists（接口 total 可能 > 单页上限 1000）"""
    all_items = []
    offset = 0
    length = 1000
    while True:
        res = call_mcp_tool("query_product_performance_asin_lists", {
            "sids": sids, "start_date": start_date, "end_date": end_date,
            "currency_code": "USD", "summary_field": "asin",
            "offset": offset, "length": length})
        items = res.get("data", {}).get("data", {}).get("list", [])
        total = res.get("data", {}).get("data", {}).get("total") or 0
        all_items.extend(items)
        if not items or len(all_items) >= total:
            break
        offset += length
        if offset > total:
            break
    return all_items


def fetch_settled_gross_profit(start_date, end_date):
    """拉订单利润报表（结算口径 gross_profit），筛南京组 15 店，按负责人拆分 Σ 结算毛利（未×0.6）。

    口径说明（2026-08-09 用户确认）：月实际毛利改用**结算毛利**（与领星前台一致），
    但仍 ×0.6 打折（采购与头程成本未完全扣减）。
    接口 query_order_profit_list 返回全组织数据，需按记录 sids 与南京组交集筛选。
    返回 dict: {"group": Σ, "hyb": 化一博Σ}（×0.6 由调用方处理）。
    """
    res = call_mcp_tool("query_order_profit_list", {
        "currency_type": "USD", "date_summary_type": 3,
        "start_date": start_date, "end_date": end_date,
        "summary_field": "parent_asin", "turn_on_summary": "1",
        "search_type": 2, "service_type": 1, "sort_type": "desc",
        "source_service": "mcp", "external_service_mark": 1,
        "length": 200, "offset": 0})
    # 兼容两种返回结构（直连一级 data / 代理两级 data）
    d = res.get("data", {}) if isinstance(res, dict) else {}
    if isinstance(d, dict):
        items = (d.get("data") or {}).get("list", []) or d.get("list", [])
    nj_set = set(str(s) for s in SIDS_LIST)
    total_g, total_hyb = 0.0, 0.0
    for r in items:
        sids = set(str(s) for s in (r.get("sids") or []))
        if sids & nj_set:
            g = float(r.get("gross_profit") or 0.0)
            total_g += g
            if "化一博" in (r.get("principal_names") or []):
                total_hyb += g
    print(f"💰 结算毛利（利润报表, {start_date}~{end_date}）: 全组 Σ gross = {total_g:.2f} (化一博 {total_hyb:.2f}) → ×0.6 = {total_g * 0.6:.2f}")
    return {"group": total_g, "hyb": total_hyb}


# 抓取并汇总领星数据
def pull_performance_and_stock(date_info):
    start_date = date_info["week_start"]
    end_date = date_info["week_end"]
    month_start = date_info["month_start"]
    month_end = date_info["month_end"]
    
    print(f"📊 正在从领星 MCP 抓取单周业绩数据 ({start_date} ~ {end_date})...")
    items = fetch_perf_all(SIDS_STR, start_date, end_date)
    
    stats = {
        "group": {"volume": 0, "amount": 0.0, "spend": 0.0, "predict_profit": 0.0, "gross_profit": 0.0, "ad_orders": 0.0, "fba_stock": 0},
        "wy":    {"volume": 0, "amount": 0.0, "spend": 0.0, "predict_profit": 0.0, "gross_profit": 0.0, "ad_orders": 0.0, "fba_stock": 0},
        "hyb":   {"volume": 0, "amount": 0.0, "spend": 0.0, "predict_profit": 0.0, "gross_profit": 0.0, "ad_orders": 0.0, "fba_stock": 0}
    }

    asin_sales_map = {}

    for item in items:
        vol = float(item.get("volume") or 0)
        amt = float(item.get("amount") or 0.0)
        spd = abs(float(item.get("spend") or 0.0))
        p_profit = float(item.get("predict_gross_profit") or 0.0)
        g_profit = float(item.get("gross_profit") or 0.0)
        ad_ord = float(item.get("ad_order_quantity") or 0)
        principals = item.get("principal_names") or []
        asin = item.get("asin")

        if asin:
            asin_sales_map[asin] = {
                "title": item.get("product_name") or item.get("title") or asin,
                "vol": vol,
                "principals": principals  # 负责人（供库存无主反查）
            }

        stats["group"]["volume"] += vol
        stats["group"]["amount"] += amt
        stats["group"]["spend"] += spd
        stats["group"]["predict_profit"] += p_profit
        stats["group"]["gross_profit"] += g_profit
        stats["group"]["ad_orders"] += ad_ord

        if "王祎" in principals:
            stats["wy"]["volume"] += vol
            stats["wy"]["amount"] += amt
            stats["wy"]["spend"] += spd
            stats["wy"]["predict_profit"] += p_profit
            stats["wy"]["gross_profit"] += g_profit
            stats["wy"]["ad_orders"] += ad_ord
        if "化一博" in principals:
            stats["hyb"]["volume"] += vol
            stats["hyb"]["amount"] += amt
            stats["hyb"]["spend"] += spd
            stats["hyb"]["predict_profit"] += p_profit
            stats["hyb"]["gross_profit"] += g_profit
            stats["hyb"]["ad_orders"] += ad_ord

    # 乘以 0.6 扣减采购与头程成本
    stats["group"]["predict_profit"] *= 0.6
    stats["wy"]["predict_profit"] *= 0.6
    stats["hyb"]["predict_profit"] *= 0.6

    # 抓取月度业绩
    print(f"📅 正在从领星 MCP 抓取当月业绩 ({month_start} ~ {month_end})...")
    month_items = fetch_perf_all(SIDS_STR, month_start, month_end)

    m_stats = {
        "group": {"amount": 0.0, "spend": 0.0, "predict_profit": 0.0, "gross_profit": 0.0},
        "wy":    {"amount": 0.0, "spend": 0.0, "predict_profit": 0.0, "gross_profit": 0.0},
        "hyb":   {"amount": 0.0, "spend": 0.0, "predict_profit": 0.0, "gross_profit": 0.0}
    }

    for item in month_items:
        amt = float(item.get("amount") or 0.0)
        spd = abs(float(item.get("spend") or 0.0))
        p_profit = float(item.get("predict_gross_profit") or 0.0)
        g_profit = float(item.get("gross_profit") or 0.0)   # 产品表现 purchase 口径结算毛利（未×0.6）
        principals = item.get("principal_names") or []

        m_stats["group"]["amount"] += amt
        m_stats["group"]["spend"] += spd
        m_stats["group"]["predict_profit"] += p_profit
        m_stats["group"]["gross_profit"] += g_profit

        if "王祎" in principals:
            m_stats["wy"]["amount"] += amt
            m_stats["wy"]["spend"] += spd
            m_stats["wy"]["predict_profit"] += p_profit
            m_stats["wy"]["gross_profit"] += g_profit
        if "化一博" in principals:
            m_stats["hyb"]["amount"] += amt
            m_stats["hyb"]["spend"] += spd
            m_stats["hyb"]["predict_profit"] += p_profit
            m_stats["hyb"]["gross_profit"] += g_profit

    # 乘以 0.6 扣减采购与头程成本
    m_stats["group"]["predict_profit"] *= 0.6
    m_stats["wy"]["predict_profit"] *= 0.6
    m_stats["hyb"]["predict_profit"] *= 0.6

    # 抓取 FBA 库存与库龄段分布 (并发多线程)
    print("📦 正在并发抓取 15 个店铺的 FBA 库存与库龄...")
    stock_stats = {
        "group": {"total_fba": 0, "a90": 0, "a181": 0, "a271": 0, "a365": 0},
        "wy":    {"total_fba": 0, "a90": 0, "a181": 0, "a271": 0, "a365": 0},
        "hyb":   {"total_fba": 0, "a90": 0, "a181": 0, "a271": 0, "a365": 0}
    }
    
    over_180_skus = set()
    over_180_qty = 0
    over_180_detail = {}   # sku -> {"name": 中文品名, "qty": 超180天数量}（2026-08-09 用户要求列出中文品名）
    high_ratio_list = []
    # 跨店去重：同一 seller_sku 在多个 sid 返回的是同一批货（欧盟泛欧9仓镜像），只计一次。
    # 注意：泛欧 SKU 各店返回的 asin_principal_list 可能不一致（如 DE 店有负责人、其他店为空），
    # 统一收集后由主线程合并，取"负责人非空"的记录。
    sku_holder = {}   # sku -> {"fba":..., "a90":..., "a181":..., "a271":..., "a365":..., "principals":[...], "asin":...}
    asin_name_map = {}  # asin -> 中文品名（全局收集，主线程合并后兜底补名，2026-08-09）
    _dedup_lock = threading.Lock()

    def fetch_single_sid_stock(sid):
        try:
            res = call_mcp_tool("get_fba_stock_list", {"sid": sid, "length": 2000})
            items = res.get("data", {}).get("list", [])
            for item in items:
                sku = item.get("seller_sku")
                fulfillable = int(item.get("afn_fulfillable_quantity") or 0)
                reserved_transfers = int(item.get("reserved_fc_transfers") or 0)
                reserved_processing = int(item.get("reserved_fc_processing") or 0)
                # 口径（2026-08-09修正）：FBA在库可售 + 待调仓 + 调仓中
                # 注意：勿用 afn_reserved_quantity（聚合预留，与 transfers/processing 重复/重叠）
                item_fba = fulfillable + reserved_transfers + reserved_processing

                a90 = int(item.get("inv_age_91_to_180_days") or 0)
                a181 = int(item.get("inv_age_181_to_270_days") or 0)
                a271 = int(item.get("inv_age_271_to_365_days") or 0)
                a365 = int(item.get("inv_age_365_plus_days") or 0)

                asin = item.get("asin")
                p_info = asin_sales_map.get(asin, {})
                # 品名只信库存行 product_name（p_info.title 可能兜底为 ASIN，不能用作品名）
                item_name = item.get("product_name") or ""

                # 负责人归属（2026-08-09修正）：库存字段 asin_principal_list 为空时，
                # 用产品表现接口的 principal_names 反查（asin 维度）
                principals = item.get("asin_principal_list") or []
                if not principals:
                    principals = p_info.get("principals") or []

                # 品名补全（与库存量无关：泛欧店 fba>0 但 product_name 为空，主店 fba=0 但品名有值）
                if sku and item_name:
                    with _dedup_lock:
                        if sku in sku_holder and not sku_holder[sku].get("name"):
                            sku_holder[sku]["name"] = item_name
                        # asin→品名全局收集（仅真实品名，避免 ASIN 污染）
                        if asin and asin not in asin_name_map:
                            asin_name_map[asin] = item_name

                # 统一收集到 sku_holder（跨店去重 + 品名补全由主线程完成）
                if sku and item_fba > 0:
                    with _dedup_lock:
                        if sku in sku_holder:
                            holder = sku_holder[sku]
                            # 已有记录负责人为空而本次非空 → 补充负责人
                            if not holder["principals"] and principals:
                                holder["principals"] = principals
                            continue
                        sku_holder[sku] = {
                            "fba": item_fba, "a90": a90, "a181": a181, "a271": a271, "a365": a365,
                            "principals": principals, "asin": asin, "name": item_name
                        }
        except Exception as e:
            print(f"Error fetching SID {sid} FBA Stock: {e}")
        return

    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = [executor.submit(fetch_single_sid_stock, sid) for sid in SIDS_LIST]
        for f in futures:
            f.result()  # 统计全部在主线程合并（sku_holder）完成

    # 主线程统一合并统计（跨店去重 + 负责人归属 + 品名补全后生成文案）
    for sku, holder in sku_holder.items():
        fba = holder["fba"]
        a90, a181, a271, a365 = holder["a90"], holder["a181"], holder["a271"], holder["a365"]
        principals = holder["principals"]
        asin = holder.get("asin", "")
        name = holder.get("name", "") or asin_name_map.get(asin, "") or asin or sku
        stock_stats["group"]["total_fba"] += fba
        stock_stats["group"]["a90"] += a90
        stock_stats["group"]["a181"] += a181
        stock_stats["group"]["a271"] += a271
        stock_stats["group"]["a365"] += a365
        if "化一博" in principals:
            stock_stats["hyb"]["total_fba"] += fba
            stock_stats["hyb"]["a90"] += a90
            stock_stats["hyb"]["a181"] += a181
            stock_stats["hyb"]["a271"] += a271
            stock_stats["hyb"]["a365"] += a365
        if "王祎" in principals:
            stock_stats["wy"]["total_fba"] += fba
            stock_stats["wy"]["a90"] += a90
            stock_stats["wy"]["a181"] += a181
            stock_stats["wy"]["a271"] += a271
            stock_stats["wy"]["a365"] += a365

        # 超过180天（2026-08-09 用户要求：列出中文品名）
        if (a181 + a271 + a365) > 0:
            over_180_qty += (a181 + a271 + a365)
            over_180_detail[sku] = {"name": name or asin or sku, "qty": a181 + a271 + a365}

        # 老品库销比>2个月（2026-08-09 用户要求：只列基础品名 + MSKU，不展示库销比）
        p_info = asin_sales_map.get(asin, {})
        vol = p_info.get("vol", 0)
        daily_avg = vol / 7.0 if vol > 0 else 0.1
        ratio = (fba / daily_avg) / 30.0
        if ratio > 2.0 and fba > 20:
            high_ratio_list.append(f"{name or asin or sku} ({sku})")

    # 组装最终结果
    res_merged = {}
    for key in ["group", "wy", "hyb"]:
        vol = stats[key]["volume"]
        amt = stats[key]["amount"]
        spd = stats[key]["spend"]
        p_profit = stats[key]["predict_profit"]
        g_profit = stats[key]["gross_profit"]
        ad_ord = stats[key]["ad_orders"]
        acoas = (spd / amt * 100) if amt > 0 else 0.0

        m_amt = m_stats[key]["amount"]
        m_spd = m_stats[key]["spend"]
        m_p_profit = m_stats[key]["predict_profit"]
        m_gross = m_stats[key]["gross_profit"]   # 产品表现 purchase 口径月结算毛利（未×0.6）
        m_acoas = (m_spd / m_amt * 100) if m_amt > 0 else 0.0

        fba_total = stock_stats[key]["total_fba"]
        daily_avg_7d = vol / 7.0 if vol > 0 else 1.0
        stock_ratio = (fba_total / daily_avg_7d) / 30.0

        res_merged[key] = {
            "volume": vol,
            "amount": amt,
            "ad_spend": spd,
            "ad_orders": ad_ord,
            "acoas": acoas,
            "predict_profit": p_profit,
            "gross_profit": g_profit,
            "month_amount": m_amt,
            "month_acoas": m_acoas,
            "month_predict_profit": m_p_profit,
            "month_perf_gross": m_gross,   # 产品表现 purchase 月口径结算毛利（未×0.6），周会纪要"结算利润额"用
            "fba_stock_total": fba_total,
            "stock_to_sales_ratio": stock_ratio,
            "age_90_180": stock_stats[key]["a90"],
            "age_181_270": stock_stats[key]["a181"],
            "age_271_365": stock_stats[key]["a271"],
            "age_365_plus": stock_stats[key]["a365"],
            "over_180_skus_cnt": len(over_180_detail),
            "over_180_qty": over_180_qty,
            "over_180_list": [f"{d['name'].split(',')[0]} ({sku}, {d['qty']}件)" for sku, d in sorted(over_180_detail.items(), key=lambda x: -x[1]["qty"])][:20],
            # 2026-08-09 用户确认：目前均为新品/全新品，老品库销比暂不统计，待有老品后启用
            "high_ratio_text": "暂不统计（当前均为新品/全新品，待有老品后启用）"
        }

    # 月实际毛利（2026-08-09 用户确认）：改用结算毛利 gross_profit × 0.6（与领星前台一致，
    # 采购与头程未完全扣减需打折）；全组 + 化一博两维度（周会纪要展示用，与报表一完全一致）
    try:
        settled = fetch_settled_gross_profit(month_start, month_end)
        res_merged["group"]["month_gross_profit"] = settled["group"] * 0.6
        res_merged["hyb"]["month_gross_profit"] = settled["hyb"] * 0.6
    except Exception as e:
        print(f"⚠️ 结算毛利拉取失败，回退预估口径: {e}")
        res_merged["group"]["month_gross_profit"] = res_merged["group"]["month_predict_profit"]
        res_merged["hyb"]["month_gross_profit"] = res_merged["hyb"]["month_predict_profit"]

    return res_merged

# 抓取 AI 表格重点任务
def fetch_aitable_tasks(meeting_date="2026-08-03"):
    print(f">>> 正在从 AI 表格【周重点任务 - 1.任务管理表】自动抓取王祎与化一博的任务...")
    res = run_dws(["aitable", "record", "list", "--all", "--base-id", BASE_ID, "--table-id", TABLE_ID])
    # dws -f json 返回嵌套结构 {"data": {"records": [...]}}
    records = (res or {}).get("data", {}).get("records") if isinstance(res, dict) else None
    if not records:
        return "1. 本周重点任务清理与推进", "1. 下周重点任务跟进", "1. 本周重点工作推进"

    dt = datetime.datetime.strptime(meeting_date, "%Y-%m-%d").date()
    curr_wk = dt.isocalendar()[1] # 32
    prev_wk = curr_wk - 1       # 31

    wy_w31 = []
    wy_w32 = []
    hyb_w31 = []
    hyb_w32 = []

    for r in records:
        cells = r.get("cells", {})
        task_name = cells.get("jxzfinmckxuusjowbz5ca", "")
        assignees = cells.get("sb82jyoeivzhh5is2guac", [])
        due_str = cells.get("y5s8sqzoulb4pdafd1mlo", "")
        week_str = str(cells.get("8LkwFxC", ""))
        
        if not task_name:
            continue
            
        uids = [u.get("userId") for u in assignees if isinstance(u, dict) and "userId" in u]

        # 判断周数类型
        is_curr = False
        is_prev = False

        if f"{curr_wk}" in week_str:
            is_curr = True
        elif f"{prev_wk}" in week_str:
            is_prev = True
        elif due_str:
            try:
                due_dt = datetime.datetime.fromisoformat(due_str).date()
                wk = due_dt.isocalendar()[1]
                if wk == curr_wk:
                    is_curr = True
                elif wk == prev_wk:
                    is_prev = True
            except Exception:
                pass

        if USER_ID_WY in uids:
            if is_prev: wy_w31.append(task_name)
            if is_curr: wy_w32.append(task_name)
        if USER_ID_HYB in uids:
            if is_prev: hyb_w31.append(task_name)
            if is_curr: hyb_w32.append(task_name)

    wy_b7_text = "\n".join([f"{i+1}. {t}" for i, t in enumerate(wy_w31)]) if wy_w31 else "1. 本周重点任务清理与推进"
    wy_b8_text = "\n".join([f"{i+1}. {t}" for i, t in enumerate(wy_w32)]) if wy_w32 else "1. 下周重点任务跟进"
    
    hyb_w31_text = "\n".join([f"{i+1}、{t}" for i, t in enumerate(hyb_w31)]) if hyb_w31 else "1、上周工作完成跟进"
    hyb_w32_text = "\n".join([f"{i+1}. {t}" for i, t in enumerate(hyb_w32)]) if hyb_w32 else "1. 本周重点工作推进"
    hyb_b8_text = f"上周工作内容\n{hyb_w31_text}\n本周工作内容\n{hyb_w32_text}"

    return wy_b7_text, wy_b8_text, hyb_b8_text

# Report 1: fill_weekly_report
def run_report1(res_data, date_info, dry_run=False):
    print("\n--- 🟢 执行报表一 (运营周复盘) ---")
    month_key = date_info["month_key"]
    targets_cfg = FY2026_TARGETS.get(month_key, FY2026_TARGETS["2026-07"])
    
    if not dry_run:
        print(f"📝 正在覆盖更新 Excel 文件: {EXCEL_TEMPLATE_R1}")
        wb = openpyxl.load_workbook(EXCEL_TEMPLATE_R1)
        sheet_weekly = wb["周报"] if "周报" in wb.sheetnames else wb.worksheets[0]
        sheet_clear = wb["清货进度表"] if "清货进度表" in wb.sheetnames else None

        row_mapping = {
            2: {"key": "group", "target_key": "group"},
            3: {"key": "wy",    "target_key": "wy"},
            4: {"key": "hyb",   "target_key": "hyb"}
        }

        for row_idx, cfg in row_mapping.items():
            data = res_data[cfg["key"]]
            t_sales = targets_cfg[cfg["target_key"]]["sales"]
            t_w_sales = t_sales / 4.0
            t_profit = targets_cfg[cfg["target_key"]]["profit"]

            sheet_weekly.cell(row=row_idx, column=4, value=round(data["amount"], 2))
            sheet_weekly.cell(row=row_idx, column=5, value=round(t_w_sales, 2))
            sheet_weekly.cell(row=row_idx, column=6, value=round(data["acoas"] / 100.0, 4))
            sheet_weekly.cell(row=row_idx, column=7, value=f"=D{row_idx}/E{row_idx}")

            sheet_weekly.cell(row=row_idx, column=8, value=round(t_sales, 2))
            sheet_weekly.cell(row=row_idx, column=9, value=round(data["month_amount"], 2))
            sheet_weekly.cell(row=row_idx, column=10, value=round(data["month_acoas"] / 100.0, 4))
            sheet_weekly.cell(row=row_idx, column=11, value=f"=I{row_idx}/H{row_idx}")

            sheet_weekly.cell(row=row_idx, column=12, value=round(t_profit, 2))
            sheet_weekly.cell(row=row_idx, column=13, value=round(data.get("month_gross_profit", data["month_predict_profit"]), 2))  # 2026-08-09: 结算×0.6
            sheet_weekly.cell(row=row_idx, column=14, value=f"=M{row_idx}/L{row_idx}")
            sheet_weekly.cell(row=row_idx, column=15, value=round(data["stock_to_sales_ratio"], 2))

            if sheet_clear:
                a90, a181, a271, a365 = data["age_90_180"], data["age_181_270"], data["age_271_365"], data["age_365_plus"]
                sheet_clear.cell(row=row_idx, column=4, value=a90)
                sheet_clear.cell(row=row_idx, column=5, value=round(a90 * 0.8))
                sheet_clear.cell(row=row_idx, column=6, value=f"=E{row_idx}/D{row_idx}")
                sheet_clear.cell(row=row_idx, column=7, value=a181)
                sheet_clear.cell(row=row_idx, column=8, value=round(a181 * 0.5))
                sheet_clear.cell(row=row_idx, column=9, value=f"=H{row_idx}/G{row_idx}")
                sheet_clear.cell(row=row_idx, column=10, value=a271)
                sheet_clear.cell(row=row_idx, column=11, value=round(a271 * 0.3))
                sheet_clear.cell(row=row_idx, column=12, value=f"=K{row_idx}/J{row_idx}")
                sheet_clear.cell(row=row_idx, column=13, value=a365)
                sheet_clear.cell(row=row_idx, column=14, value=round(a365 * 0.1))
                sheet_clear.cell(row=row_idx, column=15, value=f"=N{row_idx}/M{row_idx}")

        wb.save(EXCEL_TEMPLATE_R1)
        wb.close()
        print("Excel 填报完成，正在上传钉盘并提取链接...")
        res_drv = run_dws(["drive", "upload", "--file", EXCEL_TEMPLATE_R1])
        attach_info = {"spaceId": "", "fileId": "", "fileName": "", "fileSize": 0, "docUrl": ""}
        if res_drv and "result" in res_drv:
            result = res_drv["result"]
            attach_info = {
                "spaceId": str(result.get("spaceId")),
                "fileId": str(result.get("fileId")),
                "fileName": str(result.get("fileName")),
                "fileSize": int(result.get("fileSize", 13851)),
                "docUrl": str(result.get("docUrl"))
            }
            # 可编辑权限（2026-08-09修正）：钉盘文件无法设置角色权限，
            # 改为导入为在线电子表格并授予 EDITOR，链接用在线可编辑版
            _r1_base = os.path.splitext(os.path.basename(EXCEL_TEMPLATE_R1))[0]
            _online_url, _ = import_excel_as_online(EXCEL_TEMPLATE_R1, display_name=_r1_base)
            if _online_url:
                attach_info["docUrl"] = _online_url
    else:
        print("[Dry-run] 跳过 Excel 写入及钉盘上传")
        attach_info = {"spaceId": "MOCK_SPACE", "fileId": "MOCK_FILE", "fileName": "运营周会数据收集_王祎_v19_new.xlsx", "fileSize": 13851, "docUrl": "http://mock-doc-url"}

    # 抓取 AI 表格 WK 任务
    wy_w31_text, wy_w32_text, _ = fetch_aitable_tasks(meeting_date=date_info["meeting_date"])

    # 组装 Payload
    g_data = res_data["group"]
    g_targets = targets_cfg["group"]
    
    t_sales_monthly = g_targets["sales"]
    t_sales_weekly = t_sales_monthly / 4.0
    t_profit_monthly = g_targets["profit"]
    
    actual_profit = round(g_data.get("month_gross_profit", g_data["month_predict_profit"]), 2)  # 2026-08-09: 结算×0.6
    profit_ratio = round((actual_profit / t_profit_monthly * 100), 1) if t_profit_monthly > 0 else 0.0

    # 在线表格链接（doc import 生成的可编辑版）；后端日志接口不支持附件 type=9，链接贴备注
    _online_link = attach_info["docUrl"] if attach_info["docUrl"] and not str(attach_info["docUrl"]).startswith("http://mock") else ""

    payload = [
        # ⚠️ 2026-08-09 实测：日志后端 create_report 拒绝附件字段（type=9），dry-run 能过但真实提交报 PARAM_ERROR
        # → payload 不再含附件项，Excel 以在线表格链接贴入「备注」字段
        {"key": "周销量", "sort": "20", "content": str(int(g_data["volume"])), "contentType": "origin", "type": "2"},
        {"key": "本周实际销售额$", "sort": "21", "content": str(round(g_data["amount"], 2)), "contentType": "origin", "type": "2"},
        {"key": "AcoAs（%）", "sort": "22", "content": str(round(g_data["acoas"], 2)), "contentType": "origin", "type": "2"},
        {"key": "本周目标销售额$", "sort": "23", "content": str(round(t_sales_weekly, 2)), "contentType": "origin", "type": "2"},
        {"key": "下周目标销售额$", "sort": "24", "content": str(round(t_sales_weekly, 2)), "contentType": "origin", "type": "2"},
        {"key": "月目标销售额$", "sort": "25", "content": str(round(t_sales_monthly, 2)), "contentType": "origin", "type": "2"},
        {"key": "月实际销售额$", "sort": "26", "content": str(round(g_data["month_amount"], 2)), "contentType": "origin", "type": "2"},
        {"key": "月AcoAs（%）", "sort": "55", "content": str(round(g_data["month_acoas"], 2)), "contentType": "markdown", "type": "1"},
        {"key": "月目标毛利额$", "sort": "29", "content": str(round(t_profit_monthly, 2)), "contentType": "origin", "type": "2"},
        {"key": "月实际毛利额（未扣除工资、房租物业和财务成本等）$", "sort": "30", "content": str(actual_profit), "contentType": "origin", "type": "2"},
        {"key": "毛利额完成比率（%）", "sort": "31", "content": str(profit_ratio), "contentType": "origin", "type": "2"},
        {"key": "近30天库销比（FBA在库库存）", "sort": "57", "content": str(round(g_data["stock_to_sales_ratio"], 2)), "contentType": "markdown", "type": "1"},
        {"key": "当前库存数量——90-180天库龄", "sort": "35", "content": str(g_data["age_90_180"]), "contentType": "origin", "type": "2"},
        {"key": "目标存量——90-180天库龄", "sort": "47", "content": str(round(g_data["age_90_180"] * 0.8)), "contentType": "origin", "type": "2"},
        {"key": "当前库存数量——181-270天库龄", "sort": "38", "content": str(g_data["age_181_270"]), "contentType": "origin", "type": "2"},
        {"key": "目标存量——181-270天库龄", "sort": "48", "content": str(round(g_data["age_181_270"] * 0.5)), "contentType": "origin", "type": "2"},
        {"key": "当前库存数量——271-365天库龄", "sort": "41", "content": str(g_data["age_271_365"]), "contentType": "origin", "type": "2"},
        {"key": "目标存量——271-365天库龄", "sort": "49", "content": str(round(g_data["age_271_365"] * 0.3)), "contentType": "origin", "type": "2"},
        {"key": "当前库存数量——366天以上库龄", "sort": "44", "content": str(g_data["age_365_plus"]), "contentType": "origin", "type": "2"},
        {"key": "目标存量——366天以上库龄", "sort": "50", "content": str(round(g_data["age_365_plus"] * 0.1)), "contentType": "origin", "type": "2"},
        {"key": "本周清货计划是否符合预期？如否，写明原因", "sort": "51", "content": "符合预期", "contentType": "markdown", "type": "1"},
        {"key": "周未达成重要目标/存在问题/原因/办法", "sort": "0", "content": "无", "contentType": "markdown", "type": "1"},
        {"key": "老品月库销比超过2个月的产品名称/降低库销比方案/预期需要多久完成", "sort": "18", "content": g_data["high_ratio_text"] or "无", "contentType": "markdown", "type": "1"},
        {"key": "超过180天的SKU数量/以及对应的货量总和/预期多久清理完毕", "sort": "19",
         "content": f"超过180天SKU数量: {g_data['over_180_skus_cnt']} 个\n对应货量总和: {g_data['over_180_qty']} 件\n"
                    + ("明细:\n" + "\n".join(f"{i+1}. {x}" for i, x in enumerate(g_data['over_180_list'])) + "\n" if g_data['over_180_list'] else "")
                    + "预期清理完毕时间: 60天内通过降价与站外清完", "contentType": "markdown", "type": "1"},
        {"key": "本周重点工作及完成情况(事、量化、做到什么程度)", "sort": "17", "content": wy_w31_text, "contentType": "markdown", "type": "1"},
        {"key": "下周重点工作及计划（事、量化、时间/不超3项）", "sort": "1", "content": wy_w32_text, "contentType": "markdown", "type": "1"},
        # ===== 2026-08-09 补全模板全字段（对齐日志模板 36 字段；⚠️ 附件 type=9 后端不支持，payload 不提交）=====
        # 注：组长销售数据(sort=52)/组长库存数据(sort=53) 为 type=16 表格控件，**用户确认不填**
        # （Excel 附件中已含这些数据，客户端/附件可见，payload 不提交）
        {"key": "月目标毛利率（%）", "sort": "27",
         "content": str(round(t_profit_monthly / t_sales_monthly * 100, 2)) if t_sales_monthly > 0 else "0",
         "contentType": "origin", "type": "2"},
        {"key": "月实际毛利率（%）", "sort": "28",
         "content": str(round(actual_profit / g_data["month_amount"] * 100, 2)) if g_data["month_amount"] > 0 else "0",
         "contentType": "origin", "type": "2"},
        {"key": "周销量/销售额$/AcoAs/下周目标销售额$", "sort": "12", "content":
            f"周销量：{int(g_data['volume'])}件\n周销售额：${round(g_data['amount'], 2)}\n"
            f"AcoAs：{round(g_data['acoas'], 2)}%\n下周目标销售额：${round(t_sales_weekly, 2)}",
         "contentType": "markdown", "type": "1"},
        # 2026-08-09 用户确认：月汇总文本(sort=16)周复盘不填（下方已有月指标分享数据）
        {"key": "心得/建议/所需支持/其他需交流的问题", "sort": "2", "content": "无", "contentType": "markdown", "type": "1"},
        # 备注：贴在线可编辑表格链接（后端不支持附件 type=9，2026-08-09 实测）
        {"key": "备注", "sort": "3", "content": _online_link if _online_link else "无", "contentType": "markdown", "type": "1"},
    ]
    # 按 sort 升序排列（与模板字段顺序一致）
    payload.sort(key=lambda x: int(x["sort"]))

    payload_file = os.path.join(OUTPUT_DIR, "report_payload.json")
    os.makedirs(os.path.dirname(payload_file), exist_ok=True)
    with open(payload_file, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f"✅ 运营周报预填数据已生成！保存至: {payload_file}")
    if attach_info["docUrl"]:
        print(f"📎 附带最新 Excel 钉盘链接: {attach_info['docUrl']}")

    # 2026-08-09 用户要求：发出周报后，把在线表格链接转发给高应婷（已授权 EDITOR）
    if _online_link and not dry_run:
        _gy_open_id = "DT2AUV82nT7PZbqyiPfiPSdOwuyiPmR29xgn"   # 高应婷 openDingTalkId
        _gy_text = f"{date_info['week_start']}~{date_info['week_end']} 运营周复盘在线表格（已授权可编辑）：{_online_link}"
        try:
            res = run_dws(["chat", "message", "send", "--open-dingtalk-id", _gy_open_id,
                           "--title", "运营周复盘在线表格", "--text", _gy_text, "-y"])
            if res:
                print("📤 已转发在线表格链接给高应婷")
            else:
                print("⚠️ 转发高应婷失败（run_dws 无返回）")
        except Exception as e:
            print(f"⚠️ 转发高应婷异常: {e}")

# Report 2: generate_weekly_meeting_report
def run_report2(res_data, date_info, dry_run=False):
    print("\n--- 🟢 执行报表二 (六组周会会议纪要) ---")
    
    # 抓取 AI 表格重点任务
    _, _, hyb_b8_text = fetch_aitable_tasks(meeting_date=date_info["meeting_date"])
    wy_b7_text, _, _ = fetch_aitable_tasks(meeting_date=date_info["meeting_date"])

    dt = datetime.datetime.strptime(date_info["meeting_date"], "%Y-%m-%d")
    out_filename = f"六组周会会议纪要2026{dt.strftime('%m%d')}.xlsx"
    out_path = os.path.join(OUTPUT_DIR, out_filename)

    if not dry_run:
        print(f"📝 正在更新周报会议纪要 Excel 文件, 另存为: {out_path}")
        wb = openpyxl.load_workbook(EXCEL_TEMPLATE_R2)
        
        # --- Sheet 1: 会议记录 ---
        if "会议记录" in wb.sheetnames:
            ws1 = wb["会议记录"]
            ws1["H3"] = dt
            if wy_b7_text:
                ws1["B7"] = f"一. 本周重点任务：\n{wy_b7_text}"
            if hyb_b8_text:
                ws1["B8"] = hyb_b8_text

        # --- Sheet 2: 数据环比 ---
        if "数据环比" in wb.sheetnames:
            ws2 = wb["数据环比"]
            # 旧数据复制 (从 D -> C, J -> I)
            for r in range(3, 24):
                val_d = ws2.cell(row=r, column=4).value
                ws2.cell(row=r, column=3, value=val_d)
                val_j = ws2.cell(row=r, column=10).value
                ws2.cell(row=r, column=9, value=val_j)

            # 表头日期（动态：基于 date_info 推算上周与本周区间）
            cur_ws = datetime.datetime.strptime(date_info["week_start"], "%Y-%m-%d").date()
            cur_we = datetime.datetime.strptime(date_info["week_end"], "%Y-%m-%d").date()
            prev_ws = cur_ws - datetime.timedelta(days=7)
            prev_we = cur_ws - datetime.timedelta(days=1)
            def _fmt_short(d):
                return f"{d.month}.{d.day}"
            ws2["C2"] = f"{_fmt_short(prev_ws)}-{_fmt_short(prev_we)}"
            ws2["D2"] = f"{_fmt_short(cur_ws)}-{_fmt_short(cur_we)}"
            ws2["I2"] = f"{_fmt_short(prev_ws)}-{_fmt_short(prev_we)}"
            ws2["J2"] = f"{_fmt_short(cur_ws)}-{_fmt_short(cur_we)}"

            # 新周业绩与库存填充 (六组 & 化一博)
            g_data = res_data["group"]
            h_data = res_data["hyb"]

            ws2["D3"] = g_data["volume"]
            ws2["J3"] = h_data["volume"]
            ws2["D4"] = g_data["amount"]
            ws2["J4"] = h_data["amount"]
            ws2["D6"] = g_data["ad_orders"]
            ws2["J6"] = h_data["ad_orders"]
            
            # 订单利润额（2026-08-09 用户确认：与周复盘完全一致 = 利润报表结算 gross_profit × 0.6 月口径）
            ws2["D8"] = round(g_data.get("month_gross_profit", g_data["gross_profit"] * 0.6), 2)
            ws2["J8"] = round(h_data.get("month_gross_profit", h_data["gross_profit"] * 0.6), 2)
            
            # 结算利润额（2026-08-09 用户确认：= 产品表现 purchase（下单时间）口径 Σ gross_profit，月口径未×0.6）
            ws2["D10"] = round(g_data.get("month_perf_gross", 0), 2)
            ws2["J10"] = round(h_data.get("month_perf_gross", 0), 2)
            
            # 广告花费
            ws2["D12"] = g_data["ad_spend"]
            ws2["J12"] = h_data["ad_spend"]
            
            # 库存（在库可售+待调仓+调仓中）
            ws2["D14"] = g_data["fba_stock_total"]
            ws2["J14"] = h_data["fba_stock_total"]

            # 库龄
            ws2["D17"] = g_data["age_90_180"]
            ws2["J17"] = h_data["age_90_180"]
            ws2["D18"] = g_data["age_181_270"]
            ws2["J18"] = h_data["age_181_270"]
            ws2["D19"] = g_data["age_271_365"] + g_data["age_365_plus"]
            ws2["J19"] = h_data["age_271_365"] + h_data["age_365_plus"]

            # 环比公式绑定与留空
            ws2["D5"] = "=IF(D3=0, 0, D4/D3)"
            ws2["J5"] = "=IF(J3=0, 0, J4/J3)"
            ws2["D7"] = "=IF(D3=0, 0, D6/D3)"
            ws2["J7"] = "=IF(J6=0, 0, J6/J3)"
            ws2["D9"] = "=IF(D4=0, 0, D8/D4)"
            ws2["J9"] = "=IF(J4=0, 0, J8/J4)"
            ws2["D11"] = "=IF(D4=0, 0, D10/D4)"
            ws2["J11"] = "=IF(J4=0, 0, J10/J4)"
            ws2["D13"] = "=IF(D4=0, 0, D12/D4)"
            ws2["J13"] = "=IF(J4=0, 0, J12/J4)"
            ws2["D15"] = "=IF(D16=0, 0, D14/D16)"
            ws2["J15"] = "=IF(J16=0, 0, J14/J16)"
            ws2["D16"] = "=D3/7"
            ws2["J16"] = "=J3/7"

            # 资金使用率等强制留空
            for r in [20, 21, 22, 23]:
                ws2.cell(row=r, column=4).value = None
                ws2.cell(row=r, column=10).value = None

            # 环比公式
            ws2["E3"] = "=IF(C3=0, 0, (D3-C3)/C3)"
            ws2["K3"] = "=IF(I3=0, 0, (J3-I3)/I3)"
            ws2["E4"] = "=IF(C4=0, 0, (D4-C4)/C4)"
            ws2["K4"] = "=IF(I4=0, 0, (J4-I4)/I4)"
            ws2["E5"] = "=(D5-C5)/C5"
            ws2["K5"] = "=(J5-I5)/I5"

        wb.save(out_path)
        wb.close()
        
        # 导入为在线电子表格并授权（2026-08-09修正：钉盘文件无法设 EDITOR 权限）
        print("☁️ 正在导入 Excel 为在线表格并授予编辑权限...")
        doc_url = None
        _r2_base = os.path.splitext(os.path.basename(out_path))[0]
        doc_url, _ = import_excel_as_online(out_path, display_name=_r2_base)

        # 发送钉群卡片消息
        if doc_url:
            g_data = res_data["group"]
            h_data = res_data["hyb"]
            card_text = f"""# 📊 运营六组周会会议纪要 ({date_info["meeting_date"]})

📢 **最新周会会议纪要及数据环比在线表格已更新（具备编辑权限）！**

### 📈 核心指标速览
| 核心指标 | 运营六组全量 | 化一博个人 |
| --- | --- | --- |
| **销量** | {int(g_data['volume'])} 件 | {int(h_data['volume'])} 件 |
| **销售额** | {g_data['amount']:.2f} 美元 | {h_data['amount']:.2f} 美元 |
| **订单利润额 (*0.6, 结算·月口径)** | {g_data.get('month_gross_profit', g_data['gross_profit'] * 0.6):.2f} 美元 | {h_data.get('month_gross_profit', h_data['gross_profit'] * 0.6):.2f} 美元 |
| **结算利润额 (产品表现purchase)** | {g_data.get('month_perf_gross', 0):.2f} 美元 | {h_data.get('month_perf_gross', 0):.2f} 美元 |
| **广告花费** | {g_data['ad_spend']:.2f} 美元 | {h_data['ad_spend']:.2f} 美元 |
| **FBA在库库存** | {g_data['fba_stock_total']} 件 | {h_data['fba_stock_total']} 件 |

---
📎 **[👉 点击在线查看 / 编辑《六组周会会议纪要{dt.strftime('%m%d')}.xlsx》]({doc_url})**"""

            run_dws(["chat", "message", "send", "--group", CHAT_ID, "--title", "运营六组周会会议纪要已更新", "--text", card_text, "-y"])
            print("✅ 成功推送到钉群【南京欧洲站￥$€£】！")
    else:
        print("[Dry-run] 跳过会议纪要 Excel 生成、上传及发送钉群消息")

def main():
    import argparse
    parser = argparse.ArgumentParser(description="KovaScape 领星周报统一流水线 (Master Weekly Report Pipeline)")
    parser.add_argument("--report1", action="store_true", help="执行报表一：运营周复盘")
    parser.add_argument("--report2", action="store_true", help="执行报表二：六组周会纪要")
    parser.add_argument("--all", action="store_true", help="依次执行 report1 和 report2")
    parser.add_argument("--week-end", type=str, help="指定周六日期 (YYYY-MM-DD), 不传则基于今日自动寻找最近周六")
    parser.add_argument("--meeting-date", type=str, help="指定会议日期 (YYYY-MM-DD), 默认自动推算为周六的下周一")
    parser.add_argument("--dry-run", action="store_true", help="只拉取并校验数据，不写入物理文件和触发消息")
    
    args = parser.parse_args()
    
    if not (args.report1 or args.report2 or args.all):
        parser.print_help()
        sys.exit(1)
        
    # 1. 自动推算日期
    date_info = calculate_dates(custom_week_end=args.week_end, custom_meeting_date=args.meeting_date)
    print("\n📅 === 流水线日期推算结果 ===")
    print(f"周区间: {date_info['week_start']} ~ {date_info['week_end']}")
    print(f"月区间: {date_info['month_start']} ~ {date_info['month_end']}")
    print(f"会议日期: {date_info['meeting_date']}")
    print(f"当前分析月份: {date_info['month_key']}")
    print("============================\n")

    # 2. 抓取并汇总领星数据
    res_data = pull_performance_and_stock(date_info)
    
    # 3. 分发执行
    if args.report1 or args.all:
        run_report1(res_data, date_info, dry_run=args.dry_run)
        
    if args.report2 or args.all:
        run_report2(res_data, date_info, dry_run=args.dry_run)

if __name__ == "__main__":
    main()
